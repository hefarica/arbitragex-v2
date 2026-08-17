#!/usr/bin/env python3
"""Generate backend/searcher-rs/cartridges/manifests/math_map.json (RU-2).

Distills the canonical Excel catalog ArbitrageX_264_Cartridge_Math_Architecture.xlsx
into the machine-readable math manifest consumed by the cartridge runtime and pinned
by the CI contract test in backend/searcher-rs/src/cartridge/manifest_test.rs:

    math_map.json[i] = {
        "mev_id":          "MEV-01-001",                    # sheet 02 col 0
        "detector_id":     "R_CLOSED_CYCLE",                # sheet 02 col 14
        "primary_ops":     ["op_27", "op_21", ...],         # sheet 02 col 18 (op_XX tokens)
        "equation":        "Q_R(x)=...",                    # sheet 02 col 17
        "data_bindings":   "Full ordered route legs; ...",  # sheet 02 col 20
        "frontend_toggle": "strategy.01.001.enabled",       # sheet 02 col 5
        "mode":            "SHADOW"                         # sheet 01 col 23 (Modo inicial)
    }

The XLSX is the operator-local source of truth and is NEVER committed to the repo
(see .gitignore); only the distilled JSON is. Excel modes are law: 160 SHADOW /
104 PAPER. Every contract the Rust test enforces is validated here fail-fast so a
bad sheet never reaches a commit.

Usage: python scripts/gen_math_manifest.py [excel_path]
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    raise

ROOT = Path(__file__).resolve().parents[1]
OUT_FILE = (
    ROOT / "backend" / "searcher-rs" / "cartridges" / "manifests" / "math_map.json"
)
DEFAULT_XLSX = Path(
    r"C:\Users\HFRC\Downloads\ArbitrageX_264_Cartridge_Math_Architecture (2).xlsx"
)

MATH_SHEET = "02_CARTRIDGE_MATH_MAP"
MATRIX_SHEET = "01_MEV_MATRIX_1_11"

# Column indices (0-based), identical to scripts/generate_cartridges_v3.py.
C_MEV_ID = 0
C_FRONTEND_TOGGLE = 5
C_DETECTOR_ID = 14
C_EQUATION = 17
C_OPS_PRIMARY = 18
C_DATA_BINDINGS = 20
# In 01_MEV_MATRIX_1_11:
C_MODE = 23

EXPECTED_ROWS = 264
EXPECTED_SHADOW = 160
EXPECTED_PAPER = 104
EXPECTED_DETECTOR_FAMILIES = 60
VALID_MODES = ("SHADOW", "PAPER")

OP_RE = re.compile(r"op_(\d{2})")
MEV_ID_RE = re.compile(r"^MEV-\d{2}-\d{3}$")


def fail(msg: str) -> None:
    sys.exit(f"ERROR: {msg}")


def required_cell(row: tuple, idx: int, mev_id: str) -> str:
    value = row[idx]
    if value is None or str(value).strip() == "":
        fail(f"{mev_id}: empty required cell at column index {idx}")
    return str(value).strip()


def parse_primary_ops(raw: str, mev_id: str) -> list[str]:
    """Extract ordered, unique op_XX ids from the free-text primary-operators cell."""
    ops: list[str] = []
    seen: set[int] = set()
    for match in OP_RE.finditer(raw):
        n = int(match.group(1))
        if not 1 <= n <= 31:
            fail(f"{mev_id}: operator op_{n:02d} outside catalog op_01..op_31")
        if n not in seen:
            seen.add(n)
            ops.append(f"op_{n:02d}")
    if not ops:
        fail(f"{mev_id}: no op_XX token in primary operators cell: {raw!r}")
    return ops


def load_modes(wb) -> dict[str, str]:
    """MEV_ID -> Modo inicial (SHADOW|PAPER) from 01_MEV_MATRIX_1_11 col 23."""
    modes: dict[str, str] = {}
    for row in wb[MATRIX_SHEET].iter_rows(min_row=2, values_only=True):
        mev_id = row[C_MEV_ID]
        if mev_id is None or str(mev_id).strip() == "":
            continue
        mev_id = str(mev_id).strip()
        mode = str(row[C_MODE]).strip()
        if mode not in VALID_MODES:
            fail(f"{MATRIX_SHEET} {mev_id}: invalid Modo inicial {mode!r}")
        if mev_id in modes:
            fail(f"{MATRIX_SHEET}: duplicate MEV_ID {mev_id}")
        modes[mev_id] = mode
    return modes


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        fail(f"source Excel not found: {xlsx}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if MATH_SHEET not in wb.sheetnames or MATRIX_SHEET not in wb.sheetnames:
        fail(f"required sheets missing: {wb.sheetnames}")

    modes = load_modes(wb)
    entries: list[dict] = []
    seen_ids: set[str] = set()
    detectors: set[str] = set()
    for row in wb[MATH_SHEET].iter_rows(min_row=2, values_only=True):
        if row[C_MEV_ID] is None or str(row[C_MEV_ID]).strip() == "":
            continue
        mev_id = required_cell(row, C_MEV_ID, "?")
        if not MEV_ID_RE.match(mev_id):
            fail(f"malformed MEV_ID {mev_id!r} (expected MEV-XX-YYY)")
        if mev_id in seen_ids:
            fail(f"duplicate MEV_ID {mev_id}")
        seen_ids.add(mev_id)
        if mev_id not in modes:
            fail(f"{mev_id}: no Modo inicial found in {MATRIX_SHEET}")
        detector_id = required_cell(row, C_DETECTOR_ID, mev_id)
        detectors.add(detector_id)
        entries.append(
            {
                "mev_id": mev_id,
                "detector_id": detector_id,
                "primary_ops": parse_primary_ops(
                    required_cell(row, C_OPS_PRIMARY, mev_id), mev_id
                ),
                "equation": required_cell(row, C_EQUATION, mev_id),
                "data_bindings": required_cell(row, C_DATA_BINDINGS, mev_id),
                "frontend_toggle": required_cell(row, C_FRONTEND_TOGGLE, mev_id),
                "mode": modes[mev_id],
            }
        )

    # ── Contract counts: the Excel is law (264 = 160 SHADOW + 104 PAPER) ──
    n_shadow = sum(1 for e in entries if e["mode"] == "SHADOW")
    n_paper = len(entries) - n_shadow
    if len(entries) != EXPECTED_ROWS:
        fail(f"{MATH_SHEET}: {len(entries)} data rows, expected {EXPECTED_ROWS}")
    if set(modes) != seen_ids:
        only_math = seen_ids - set(modes)
        only_matrix = set(modes) - seen_ids
        fail(f"MEV_ID drift between sheets: math-only={sorted(only_math)} "
             f"matrix-only={sorted(only_matrix)}")
    if len(detectors) != EXPECTED_DETECTOR_FAMILIES:
        fail(f"{len(detectors)} detector families, expected {EXPECTED_DETECTOR_FAMILIES}")
    if (n_shadow, n_paper) != (EXPECTED_SHADOW, EXPECTED_PAPER):
        fail(f"mode split {n_shadow}/{n_paper}, expected "
             f"{EXPECTED_SHADOW}/{EXPECTED_PAPER}")

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    # Round-trip guard: the committed artifact must parse exactly as written.
    json.loads(OUT_FILE.read_text(encoding="utf-8"))
    print(
        f"math_map.json: {len(entries)} entries "
        f"({n_shadow} SHADOW / {n_paper} PAPER, "
        f"{len(detectors)} detector families) -> {OUT_FILE.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
