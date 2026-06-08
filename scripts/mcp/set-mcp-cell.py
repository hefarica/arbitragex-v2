#!/usr/bin/env python3
"""
set-mcp-cell.py — set ONE 'MCP Wiring' cell value from env vars, without echoing
the value. Preserves every other cell. Use to wire a secret/URL produced by
automation (e.g. a freshly created read-only credential) into the Excel SSOT.

Usage (value comes from env, NOT argv, to limit exposure):
  MCP_VAR=REDIS_URL_READONLY MCP_VAL='redis://...' python scripts/mcp/set-mcp-cell.py
"""
import os, sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  pip install openpyxl")

XLSX = Path(os.environ.get(
    "MCP_SECRETS_XLSX", Path.home() / "Downloads" / "ArbitrageX_Secrets_Config.xlsx"))
SHEET = "MCP Wiring"

var = (os.environ.get("MCP_VAR") or "").strip()
val = os.environ.get("MCP_VAL")
if not var or val is None:
    sys.exit("ERROR: set MCP_VAR and MCP_VAL env vars.")

lock = XLSX.with_name("~$" + XLSX.name)
if lock.exists():
    sys.exit(f"ERROR: '{XLSX.name}' is OPEN in Excel. Close it and re-run.")

wb = openpyxl.load_workbook(XLSX)
if SHEET not in wb.sheetnames:
    sys.exit(f"ERROR: no '{SHEET}' sheet. Run add-mcp-rows-to-xlsx.py first.")

ws = wb[SHEET]
for row in ws.iter_rows(min_col=1, max_col=2):
    if row[0].value and str(row[0].value).strip() == var:
        row[1].value = val
        row[1].fill = PatternFill(fill_type=None)  # clear yellow: it's filled now
        wb.save(XLSX)
        print(f"OK: set {var} (len={len(val)}) in '{SHEET}'.")
        break
else:
    sys.exit(f"ERROR: variable '{var}' row not found in '{SHEET}'.")
