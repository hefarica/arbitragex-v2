#!/usr/bin/env python3
"""
add-mcp-rows-to-xlsx.py — add an 'MCP Wiring' sheet to the operator's secrets
workbook with EMPTY yellow placeholder rows for the MCP vars not yet present,
so the operator only pastes values. Then `gen-env-mcp.py` picks them up.

Safe & idempotent:
  - Backs up the workbook first (timestamped copy).
  - Only ADDS/refreshes the 'MCP Wiring' sheet; never touches other sheets.
  - Leaves 'Valor' cells EMPTY (the generator skips empties), so adding rows
    can never wire a bogus value.
  - Never reads or prints existing secret values.

Run:  python scripts/mcp/add-mcp-rows-to-xlsx.py
"""
import os, shutil, sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("ERROR: openpyxl not installed.  pip install openpyxl")

XLSX = Path(os.environ.get(
    "MCP_SECRETS_XLSX",
    Path.home() / "Downloads" / "ArbitrageX_Secrets_Config.xlsx"))
SHEET = "MCP Wiring"

# (Variable in col A, Notas in col C, default for col B). Empty default -> yellow
# placeholder the generator skips until you paste a value.
ROWS = [
    ("GITHUB_TOKEN",
     "PAT de GitHub READ-ONLY (fine-grained, contents:read). Sin scopes de escritura/admin.", ""),
    ("MAGIC_21ST_API_KEY",
     "API key de 21st.dev Magic (generador de componentes UI).", ""),
    ("CONTEXT7_API_KEY",
     "Opcional: context7 ya funciona via plugin. Solo si quieres una copia user-scope propia.", ""),
    ("EVM_NETWORK",
     "Red para construir EVM_RPC_URL desde ALCHEMY_API_KEY (eth-mainnet, base-mainnet, "
     "arb-mainnet, opt-mainnet...). NO es secreto; cambiala si quieres otra cadena.", "eth-mainnet"),
    ("EVM_RPC_URL",
     "RPC read-only. Si lo dejas vacio se construye con EVM_NETWORK + ALCHEMY_API_KEY.", ""),
    ("ANVIL_FORK_RPC_URL",
     "Fork local de Anvil (solo lectura/simulacion). NUNCA private key: PRIVATE_KEY queda vacio.", ""),
    ("REDIS_URL_READONLY",
     "Override ACL read-only: redis://arbx_ro:PASS@HOST:PORT (ACL +@read -@write -@dangerous).", ""),
    ("DATABASE_URL_READONLY",
     "Opcional override. Si lo dejas vacio se construye desde ARBX_RO_PASSWORD (rol arbx_ro).", ""),
]


def main():
    if not XLSX.exists():
        sys.exit(f"ERROR: secrets file not found: {XLSX}")

    # Refuse early if Excel has the file open (lock file ~$<name>) so we don't
    # build anything we can't save.
    lock = XLSX.with_name("~$" + XLSX.name)
    if lock.exists():
        sys.exit(f"ERROR: '{XLSX.name}' is OPEN in Excel (lock {lock.name}). "
                 f"Close Excel and re-run — original untouched, no backup created.")

    try:
        wb = openpyxl.load_workbook(XLSX)  # default: keeps formulas + formatting
    except Exception as e:
        sys.exit(f"ERROR loading workbook (is it open in Excel? close it): {e}")

    prior = {}
    if SHEET in wb.sheetnames:
        for r in wb[SHEET].iter_rows(values_only=True):
            if r and len(r) >= 2 and r[0] and r[1] not in (None, ""):
                prior[str(r[0]).strip()] = r[1]   # preserve already-typed values
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    yellow = PatternFill("solid", fgColor="FFFF00")
    grey = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "MCP Wiring — llena SOLO la columna 'Valor' (amarilla). Deja vacio lo que no uses."
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:C1")
    ws["A2"] = ("Luego corre:  python scripts/mcp/gen-env-mcp.py   "
                "(no se hardcodea nada: el generador lee estas celdas hacia .env.mcp)")
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells("A2:C2")

    hdr = 4
    for col, name in enumerate(["Variable", "Valor", "Notas"], 1):
        c = ws.cell(row=hdr, column=col, value=name)
        c.font, c.fill, c.border = Font(bold=True), grey, border

    for i, (var, note, default) in enumerate(ROWS, start=hdr + 1):
        val = prior.get(var, default)              # keep operator-typed value; else default
        a = ws.cell(row=i, column=1, value=var)
        a.font, a.border = Font(bold=True), border
        b = ws.cell(row=i, column=2, value=(val if val != "" else None))
        b.border = border
        if val == "":
            b.fill = yellow                        # empty -> highlight; generator skips it
        c = ws.cell(row=i, column=3, value=note)
        c.border, c.alignment = border, Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 72

    bak = XLSX.with_name(f"{XLSX.stem}.backup-{datetime.now():%Y%m%d-%H%M%S}{XLSX.suffix}")
    shutil.copy2(XLSX, bak)
    try:
        wb.save(XLSX)
    except PermissionError:
        try:
            bak.unlink()  # save never happened — drop the redundant backup
        except OSError:
            pass
        sys.exit("ERROR: cannot save — Excel has the file open. Close it and re-run. "
                 "Original is intact.")

    print(f"OK. Added '{SHEET}' with {len(ROWS)} empty yellow placeholder rows.")
    print(f"Backup: {bak}")
    print("Next: paste values into the yellow cells, then run:")
    print("  python scripts/mcp/gen-env-mcp.py")


if __name__ == "__main__":
    main()
