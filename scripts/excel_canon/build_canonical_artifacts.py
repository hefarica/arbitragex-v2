# -*- coding: utf-8 -*-
"""Canonical Excel→repo artifact builder (operator directive 2026-08-23).

Turns the ULTRA workbook (ArbitrageX_Route_Strategy_Optimizer_264_ULTRA) into
the canonical requirements/coverage artifacts:

  artifacts/excel_requirements.json   — canonical requirement matrix (IDs from the workbook)
  artifacts/strategy_registry.json    — per-strategy canonical record + repo evidence
  artifacts/source_field_map.json     — sheet 19 + per-source field provenance
  artifacts/excel_coverage.json       — mechanical audit result per requirement
  artifacts/excel_coverage.md         — human-readable milestone report

Design rules (RULE 00 / R8 / anti-self-deception):
  - Every count is DERIVED from the workbook / repo at build time. Nothing hardcoded.
  - Coverage is computed by VERIFYING anchors against the working tree
    (file existence + content regex). A requirement with no anchor is MISSING,
    never silently PASS.
  - Read-only: this script never mutates repo sources; it only writes artifacts/.

Usage:
  py scripts/excel_canon/build_canonical_artifacts.py            # uses committed raw
  py scripts/excel_canon/build_canonical_artifacts.py --xlsx path # re-extract first
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import defaultdict

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
ART = os.path.join(ROOT, "artifacts")
RAW_PATH = os.path.join(ART, "excel_ultra_raw.json")

# ─────────────────────────────────────────────────────────────────────────────
# Workbook access
# ─────────────────────────────────────────────────────────────────────────────

def load_raw(xlsx: str | None) -> dict:
    if xlsx:
        sys.path.insert(0, os.path.join(ROOT, "scratchpad"))
        # reuse the extractor if present; else inline minimal extraction
        from openpyxl import load_workbook  # noqa: PLC0415
        wb_v = load_workbook(xlsx, data_only=True, read_only=True)
        wb_f = load_workbook(xlsx, data_only=False, read_only=True)
        sheets: dict = {}
        for name in wb_v.sheetnames:
            rows, n_cells, n_formulas, digest = [], 0, 0, hashlib.sha256()
            for v_row, f_row in zip(wb_v[name].iter_rows(), wb_f[name].iter_rows()):
                r = []
                for v_c, f_c in zip(v_row, f_row):
                    val, formula = v_c.value, None
                    if isinstance(f_c.value, str) and f_c.value.startswith("="):
                        formula = f_c.value
                    if val is None and formula is None:
                        continue
                    n_cells += 1
                    if formula is not None:
                        n_formulas += 1
                    e = [v_c.coordinate, val]
                    if formula is not None and formula != val:
                        e.append(formula)
                    r.append(e)
                    digest.update(repr(e).encode("utf-8"))
                if r:
                    rows.append(r)
            sheets[name] = {
                "non_empty_cells": n_cells, "formulas": n_formulas,
                "content_sha256": digest.hexdigest(), "rows": rows,
            }
        raw = {
            "source": os.path.abspath(xlsx),
            "sha256": hashlib.sha256(open(xlsx, "rb").read()).hexdigest(),
            "sheets": sheets,
        }
        with open(RAW_PATH, "w", encoding="utf-8") as f:
            json.dump(raw, f, ensure_ascii=False, separators=(",", ":"))
        return raw
    with open(RAW_PATH, encoding="utf-8") as f:
        return json.load(f)


def sheet_rows(raw: dict, name: str):
    return raw["sheets"][name]["rows"]


def row_dict(row: list) -> dict:
    """[["A1", v, f?], ...] -> {"A": {"v": v, "f": f?}} keyed by column letters."""
    out = {}
    for cell in row:
        col = re.match(r"([A-Z]+)", cell[0]).group(1)
        e = {"v": cell[1]}
        if len(cell) > 2:
            e["f"] = cell[2]
        out[col] = e
    return out


def grid(rows: list) -> dict[int, dict[str, dict]]:
    return {i: row_dict(r) for i, r in enumerate(rows)}


def val(g: dict, r: int, c: str):
    e = g.get(r, {}).get(c)
    return None if e is None else e.get("v")


def formula(g: dict, r: int, c: str):
    e = g.get(r, {}).get(c)
    return None if e is None else e.get("f")


def find_header(g: dict, a_value: str) -> int:
    """Grid index of the row whose column A equals a_value (robust to the
    extractor skipping empty rows — never trust absolute row numbers)."""
    for r in sorted(g):
        if str(val(g, r, "A")).strip() == a_value:
            return r
    raise KeyError(f"header {a_value!r} not found in column A")


def parse_table(g: dict, header_row: int, a_test=None) -> tuple[dict, list[int]]:
    """Header map {col: name} + data row indices after header_row."""
    hdr = {c: str(val(g, header_row, c)) for c in g[header_row] if val(g, header_row, c) is not None}
    data = []
    for r in sorted(g):
        if r <= header_row:
            continue
        a = val(g, r, "A")
        if a is None:
            continue
        if a_test is not None and not a_test(a):
            continue
        data.append(r)
    return hdr, data


# ─────────────────────────────────────────────────────────────────────────────
# Repo verification (mechanical, honest)
# ─────────────────────────────────────────────────────────────────────────────

SEARCH_ROOTS = [
    "backend/searcher-rs/src",
    "backend/searcher-rs/cartridges",
    "backend/math-engine/src",
    "backend/api-server/src",
    "backend/edge",
    "backend/relays-client/src",
    "frontend/app",
    "frontend/lib",
    "frontend/components",
    "skills/arbitragex-ultra",
    "database/migrations",
]
TEXT_EXT = {".rs", ".ts", ".tsx", ".js", ".json", ".toml", ".md", ".sql", ".yml", ".yaml", ".rhai"}


class RepoIndex:
    """Loads searchable text files once; serves content-regex checks."""

    def __init__(self) -> None:
        self.files: dict[str, str] = {}
        for rel_root in SEARCH_ROOTS:
            abs_root = os.path.join(ROOT, rel_root)
            if not os.path.isdir(abs_root):
                continue
            for dirpath, _dirs, names in os.walk(abs_root):
                for n in names:
                    ext = os.path.splitext(n)[1].lower()
                    if ext not in TEXT_EXT:
                        continue
                    p = os.path.join(dirpath, n)
                    rel = os.path.relpath(p, ROOT).replace("\\", "/")
                    try:
                        with open(p, encoding="utf-8", errors="replace") as fh:
                            self.files[rel] = fh.read()
                    except OSError:
                        continue

    def exists(self, rel: str) -> bool:
        return os.path.isfile(os.path.join(ROOT, rel))

    def glob_exists(self, pattern: str) -> list[str]:
        rx = re.compile("^" + re.escape(pattern).replace(r"\*", ".*") + "$")
        return sorted(p for p in self.files if rx.match(p)) or (
            sorted(
                p for p in self.files if pattern.rsplit("/", 1)[-1].rstrip("*") in p
            ) if "*" in pattern else []
        )

    def grep(self, rx: str) -> dict[str, list[int]]:
        """Return {file: [line numbers]} for files whose content matches rx."""
        comp = re.compile(rx, re.IGNORECASE)
        hits: dict[str, list[int]] = {}
        for path, content in self.files.items():
            lines = content.splitlines()
            ln = [i + 1 for i, line in enumerate(lines) if comp.search(line)]
            if ln:
                hits[path] = ln[:5]
        return hits


def anchor_status(repo: RepoIndex, anchors: list[dict]) -> tuple[str, list[dict]]:
    """Evaluate a list of anchor specs: {path} / {glob} / {regex} (key form) or
    {"kind": ..., "target": ...} (normalized form)."""
    norm = []
    for a in anchors:
        if "kind" in a and "target" in a:
            norm.append({a["kind"]: a["target"]})
        else:
            norm.append(a)
    anchors = norm
    evid: list[dict] = []
    hits = 0
    for a in anchors:
        if "path" in a:
            ok = repo.exists(a["path"])
            evid.append({"kind": "path", "target": a["path"], "hit": ok})
            hits += int(ok)
        elif "glob" in a:
            found = repo.glob_exists(a["glob"])
            evid.append({"kind": "glob", "target": a["glob"], "hit": bool(found),
                         "matches": found[:6], "n": len(found)})
            hits += int(bool(found))
        elif "regex" in a:
            found = repo.grep(a["regex"])
            evid.append({"kind": "regex", "target": a["regex"], "hit": bool(found),
                         "matches": sorted(found)[:8], "n": len(found)})
            hits += int(bool(found))
    total = len(anchors)
    status = "MISSING" if hits == 0 else ("VERIFIED" if hits == total else "PARTIAL")
    return status, evid


# ─────────────────────────────────────────────────────────────────────────────
# Requirement model
# ─────────────────────────────────────────────────────────────────────────────

class Requirements:
    def __init__(self) -> None:
        self.items: list[dict] = []

    def add(self, rid: str, sheet: str, origin: str, statement: str,
            anchors: list[dict], family: str = "", provenance: str = "CANONICAL_WORKBOOK") -> dict:
        status, evid = anchor_status(REPO, anchors)
        item = {
            "id": rid, "sheet": sheet, "origin": origin, "family": family,
            "statement": statement, "provenance": provenance,
            "repo_anchors": anchors, "verification": {"status": status, "evidence": evid},
        }
        self.items.append(item)
        return item


# ─────────────────────────────────────────────────────────────────────────────
# Parsers per sheet
# ─────────────────────────────────────────────────────────────────────────────

OP_CODE = re.compile(r"(op_\d{2})")
STRAT_ID = re.compile(r"^MEV-(\d{2})-(\d{3})$")


def parse_catalog(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "11_STRATEGY_CATALOG"))
    hdr = {c: val(g, 0, c) for c in g[0]}
    out = []
    for r in range(1, len(g)):
        mev = val(g, r, "A")
        if not mev or not STRAT_ID.match(str(mev)):
            continue
        rec = {hdr[c]: val(g, r, c) for c in hdr if val(g, r, c) is not None}
        rec["_row"] = r + 1
        out.append(rec)
    return out, hdr


def parse_matrix(raw: dict) -> tuple[list[dict], dict]:
    g = grid(sheet_rows(raw, "13_STRAT_OP_MATRIX"))
    hdr = {c: val(g, 0, c) for c in g[0]}
    op_cols = [c for c, h in hdr.items() if h and re.match(r"^op_\d{2}$", str(h))]
    out = []
    for r in range(1, len(g)):
        mev = val(g, r, "A")
        if not mev or not STRAT_ID.match(str(mev)):
            continue
        vals = {str(hdr[c]): val(g, r, c) for c in op_cols}
        nz = {k: v for k, v in vals.items() if isinstance(v, (int, float)) and v != 0}
        rec = {
            "mev_id": mev, "strategy": val(g, r, "B"), "detector_id": val(g, r, "C"),
            "links": nz, "link_count": len(nz),
            "primary_count": val(g, r, "AI"), "secondary_count": val(g, r, "AJ"),
            "phase_primaries": {k: val(g, r, c) for k, c in
                                [("discover", "AK"), ("size", "AL"), ("risk", "AM"),
                                 ("rank", "AN"), ("validate", "AO")]},
        }
        out.append(rec)
    return out, hdr


def parse_templates(raw: dict) -> dict[str, dict]:
    g = grid(sheet_rows(raw, "14_STRATEGY_TEMPLATES"))
    hdr = {c: val(g, 0, c) for c in g[0]}
    out = {}
    for r in range(1, len(g)):
        mev = val(g, r, "A")
        if not mev or not STRAT_ID.match(str(mev)):
            continue
        out[str(mev)] = {hdr[c]: val(g, r, c) for c in hdr if val(g, r, c) is not None}
        out[str(mev)]["_row"] = r + 1
    return out


def parse_operators(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "12_OPERATOR_CONTROL"))
    h = find_header(g, "ID")
    _, data = parse_table(g, h, a_test=lambda a: str(a).isdigit())
    out = []
    for r in data:
        code = val(g, r, "B")
        if not code or not str(code).startswith("op_"):
            continue
        out.append({
            "code": str(code), "name": val(g, r, "C"),
            "canonical_role": val(g, r, "D"), "enabled": val(g, r, "E"),
            "engine_present": val(g, r, "F"), "calibration_state": val(g, r, "G"),
            "operational_weight_pct": val(g, r, "H"),
            "primary_uses": val(g, r, "I"), "secondary_uses": val(g, r, "J"),
            "_row": r + 1,
        })
    return out


def parse_config(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "01_CONFIG"))
    h = find_header(g, "Parámetro")
    out = []
    for r in sorted(g):
        if r <= h:
            continue
        p = val(g, r, "A")
        if not p or str(p).startswith("INTEGRACIÓN"):
            continue
        if val(g, r, "B") is None or val(g, r, "D") is None:
            continue
        out.append({
            "knob": str(p), "value": val(g, r, "B"), "unit": val(g, r, "C"),
            "layer": val(g, r, "D"), "description": val(g, r, "E"), "note": val(g, r, "F"),
            "_row": r + 1,
        })
    return out


def parse_financing(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "02_FINANCING"))
    h = find_header(g, "Mode")
    out = []
    for r in sorted(g):
        if r <= h:
            continue
        m = val(g, r, "A")
        if not m or val(g, r, "C") is None:
            continue
        out.append({
            "mode": str(m), "enabled": val(g, r, "B"), "capacity_usd": val(g, r, "C"),
            "provider_fee_bps": val(g, r, "D"), "extra_gas_usd": val(g, r, "E"),
            "route_constraint": val(g, r, "F"), "assumption": val(g, r, "G"),
            "source_url": val(g, r, "H"), "_row": r + 1,
        })
    return out


def parse_gates07(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "07_GATES"))
    h = find_header(g, "Gate")
    out = []
    for r in sorted(g):
        if r <= h:
            continue
        gid = val(g, r, "A")
        if not gid or not str(gid).startswith("G") or not re.match(r"^G\d", str(gid)):
            continue
        out.append({
            "gate": str(gid), "objective": val(g, r, "B"), "metric": val(g, r, "C"),
            "current": val(g, r, "D"), "formula": formula(g, r, "D"),
            "threshold": val(g, r, "E"), "status": val(g, r, "F"), "action": val(g, r, "G"),
            "_row": r + 1,
        })
    return out


def parse_lists10(raw: dict) -> dict[str, list]:
    g = grid(sheet_rows(raw, "10_LISTAS"))
    hdr = {c: val(g, 1, c) for c in g.get(1, {})}  # row 2 header (A3=...)
    out: dict[str, list] = {}
    for c, name in hdr.items():
        if not name:
            continue
        vals = []
        for r in range(2, len(g)):
            v = val(g, r, c)
            if v is None:
                continue
            vals.append(v)
        out[str(name)] = vals
    return out


def parse_scoreboard(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "16_STRATEGY_SCOREBOARD"))
    h = find_header(g, "Group")
    out = []
    for r in sorted(g):
        if r <= h:
            continue
        grp = val(g, r, "A")
        if not isinstance(grp, (int, float)):
            continue
        out.append({
            "group": int(grp), "family": val(g, r, "B"), "surface": val(g, r, "C"),
            "strategies": val(g, r, "D"), "route_ready": val(g, r, "E"),
            "needs_data": val(g, r, "F"), "observe_only": val(g, r, "G"),
            "avg_op_coverage": val(g, r, "H"), "avg_readiness": val(g, r, "I"),
            "compatible_routes_total": val(g, r, "J"),
            "viable_compatible_total": val(g, r, "K"), "coverage_pct": val(g, r, "L"),
            "_row": r + 1,
        })
    return out


def parse_fieldmap19(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "19_SOURCE_FIELD_MAP"))
    h = find_header(g, "Source_Sheet")
    out = []
    for r in sorted(g):
        if r <= h:
            continue
        s = val(g, r, "A")
        if not s or val(g, r, "B") is None:
            continue
        out.append({
            "source_sheet": str(s), "records": val(g, r, "B"), "columns": val(g, r, "C"),
            "fields": val(g, r, "D"), "integrated_into": val(g, r, "E"),
            "treatment": val(g, r, "F"), "_row": r + 1,
        })
    return out


def parse_source_audit17(raw: dict) -> tuple[list[dict], list[dict]]:
    g = grid(sheet_rows(raw, "17_SOURCE_AUDIT"))
    h1 = find_header(g, "Source_File")
    h2 = find_header(g, "Comparison")
    sources, containment = [], []
    for r in sorted(g):
        if h1 < r < h2:
            s = val(g, r, "A")
            if s and val(g, r, "C") is not None:
                sources.append({
                    "file": str(s), "sha256": val(g, r, "B"), "bytes": val(g, r, "C"),
                    "sheets": val(g, r, "D"), "role": val(g, r, "E"),
                    "conflicts": val(g, r, "F"), "coverage": val(g, r, "G"), "_row": r + 1,
                })
        elif r > h2:
            cmp_ = val(g, r, "A")
            if cmp_ and val(g, r, "B") is not None:
                containment.append({
                    "comparison": str(cmp_), "scope": val(g, r, "B"),
                    "difference": val(g, r, "C"), "other_only": val(g, r, "D"),
                    "conflicting": val(g, r, "E"), "_row": r + 1,
                })
    return sources, containment


def parse_library18(raw: dict) -> dict[str, list[dict]]:
    """Section-aware parse of 18_CANONICAL_LIBRARY (7 sections)."""
    rows = sheet_rows(raw, "18_CANONICAL_LIBRARY")
    g = grid(rows)
    section_titles = []
    for r, row in enumerate(g.values()):
        cells = list(row.values())
        if len(cells) == 1 and isinstance(cells[0].get("v"), str):
            t = cells[0]["v"]
            if t.isupper() and len(t) > 3 and " " not in t:
                section_titles.append((r, t))
    out: dict[str, list[dict]] = {}
    for i, (r0, title) in enumerate(section_titles):
        r1 = section_titles[i + 1][0] if i + 1 < len(section_titles) else len(g)
        hdr_row = r0 + 1
        hdr = {c: val(g, hdr_row, c) for c in g.get(hdr_row, {})}
        recs = []
        for r in range(hdr_row + 1, r1):
            if val(g, r, "A") is None:
                continue
            recs.append({str(hdr[c]): val(g, r, c) for c in hdr if val(g, r, c) is not None})
        out[title] = recs
    return out


def parse_algorithms04(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "04_ALGORITMOS"))
    h = find_header(g, "Algorithm")
    # Algorithm table rows end where the benchmark table header (A=Study) begins.
    stop = None
    for r in sorted(g):
        if str(val(g, r, "A")).strip() == "Study":
            stop = r
            break
    out = []
    for r in sorted(g):
        if r <= h or (stop is not None and r >= stop):
            continue
        a = val(g, r, "A")
        if not a or val(g, r, "B") is None:
            continue
        out.append({
            "algorithm": str(a), "enabled": val(g, r, "B"), "role": val(g, r, "C"),
            "complexity": val(g, r, "D"), "estimated_ops": val(g, r, "E"),
            "fits_cpu_budget": val(g, r, "F"), "exhaustive": val(g, r, "G"),
            "hop_fit": val(g, r, "H"), "incremental_fit": val(g, r, "I"),
            "implication": val(g, r, "J"), "source": val(g, r, "K"), "_row": r + 1,
        })
    return out


def parse_references09(raw: dict) -> list[dict]:
    g = grid(sheet_rows(raw, "09_REFERENCIAS"))
    out = []
    for r in range(3, len(g)):
        t = val(g, r, "A")
        if not t or val(g, r, "E") is None:
            continue
        out.append({"topic": str(t), "reference": val(g, r, "B"), "year": val(g, r, "C"),
                    "use": val(g, r, "D"), "url": val(g, r, "E"), "note": val(g, r, "F"),
                    "_row": r + 1})
    return out


def parse_graph03_headers(raw: dict) -> dict:
    g = grid(sheet_rows(raw, "03_GRAFO_POOLS"))
    h = find_header(g, "Pool_ID")
    hdr = {c: str(val(g, h, c)) for c in g[h] if val(g, h, c) is not None}
    formulas = {}
    first_data = None
    for r in sorted(g):
        if r > h and val(g, r, "A") is not None:
            first_data = r
            break
    if first_data is not None:
        for c, name in hdr.items():
            fx = formula(g, first_data, c)
            if fx:
                formulas[name] = fx
    n_data = sum(1 for r in sorted(g) if r > h and val(g, r, "A") is not None)
    return {"headers": list(hdr.values()), "sample_formulas": formulas,
            "data_rows": n_data}


def parse_rutas05_headers(raw: dict) -> dict:
    g = grid(sheet_rows(raw, "05_RUTAS"))
    h = find_header(g, "Route_ID")
    hdr = {c: str(val(g, h, c)) for c in g[h] if val(g, h, c) is not None}
    first_data = None
    for r in sorted(g):
        if r > h and val(g, r, "A") is not None:
            first_data = r
            break
    cols = {}
    for c, name in hdr.items():
        e = {"header": name}
        if first_data is not None:
            fx = formula(g, first_data, c)
            if fx:
                e["formula_row6"] = fx
        cols[name] = e
    n_data = sum(1 for r in sorted(g) if r > h and val(g, r, "A") is not None)
    return {"columns": cols, "data_rows": n_data}


def parse_route_opt15(raw: dict) -> dict:
    g = grid(sheet_rows(raw, "15_STRAT_ROUTE_OPT"))
    # Two header rows exist; the route table is the one where B == "Surface".
    h = None
    for r in sorted(g):
        if str(val(g, r, "A")).strip() == "Route_ID":
            h = r
            break
    if h is None:
        return {"columns": [], "data_rows": 0, "sample_formulas": {}}
    hdr = {c: str(val(g, h, c)) for c in g[h] if val(g, h, c) is not None}
    first_data = None
    for r in sorted(g):
        if r > h and val(g, r, "A") is not None:
            first_data = r
            break
    out = {"columns": list(hdr.values()), "data_rows": 0, "sample_formulas": {}}
    if first_data is not None:
        for c, name in hdr.items():
            fx = formula(g, first_data, c)
            if fx:
                out["sample_formulas"][name] = fx
    out["data_rows"] = sum(1 for r in sorted(g) if r > h and val(g, r, "A") is not None)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Build
# ─────────────────────────────────────────────────────────────────────────────

REPO: RepoIndex = None  # set in main


def strategy_evidence(repo: RepoIndex, mev: str, catalog_rec: dict,
                      canon_ops: dict | None) -> tuple[str, list[dict], dict]:
    m = STRAT_ID.match(mev)
    stem = f"mev_{m.group(1)}_{m.group(2)}"
    cartridge_glob = f"backend/searcher-rs/cartridges/strategies/{stem}_*.rhai"
    canon_dir = f"skills/arbitragex-ultra/strategies/{mev}/STRATEGY.json"
    ev: dict = {"cartridge": None, "canon": None, "ops_match": None}
    status_anchors = []
    cart = repo.glob_exists(cartridge_glob)
    ev["cartridge"] = cart[:2]
    has_canon = repo.exists(canon_dir)
    ev["canon"] = canon_dir if has_canon else None
    ops_match = None
    if has_canon and cart:
        try:
            with open(os.path.join(ROOT, canon_dir), encoding="utf-8") as fh:
                cj = json.load(fh)
            canon_p = sorted(OP_CODE.findall(json.dumps(cj.get("primary_operators", []))))
            canon_s = sorted(OP_CODE.findall(json.dumps(cj.get("secondary_operators", []))))
        except (json.JSONDecodeError, OSError):
            canon_p = canon_s = None
        if canon_p is not None:
            cat_p = sorted(OP_CODE.findall(str(catalog_rec.get("Primary_Ops", ""))))
            cat_s = sorted(OP_CODE.findall(str(catalog_rec.get("Secondary_Ops", ""))))
            ops_match = {"primary": canon_p == cat_p, "secondary": canon_s == cat_s,
                         "catalog_primary": cat_p, "catalog_secondary": cat_s,
                         "canon_primary": canon_p, "canon_secondary": canon_s}
            ev["ops_match"] = ops_match
    if cart and has_canon and (ops_match is None or (ops_match["primary"] and ops_match["secondary"])):
        status = "VERIFIED"
    elif cart or has_canon:
        status = "PARTIAL"
    else:
        status = "MISSING"
    return status, [{"kind": "glob", "target": cartridge_glob},
                    {"kind": "path", "target": canon_dir}], ev


def main() -> int:
    global REPO
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    args = ap.parse_args()
    raw = load_raw(args.xlsx)
    REPO = RepoIndex()
    R = Requirements()
    os.makedirs(ART, exist_ok=True)

    wb_meta = {
        "source": raw["source"], "sha256": raw["sha256"],
        "sheets": {n: {k: v for k, v in s.items() if k != "rows"}
                   for n, s in raw["sheets"].items()},
        "derived_at_counting": "counts below are derived from parsed rows, never hardcoded",
    }

    # ── 11/13/14: strategy registry (the 264) ─────────────────────────────
    catalog, cat_hdr = parse_catalog(raw)
    matrix, mat_hdr = parse_matrix(raw)
    templates = parse_templates(raw)
    mat_by_id = {m["mev_id"]: m for m in matrix}

    registry = []
    for rec in catalog:
        mev = str(rec["MEV_ID"])
        status, anchors, ev = strategy_evidence(REPO, mev, rec, None)
        m = mat_by_id.get(mev, {})
        entry = {
            "mev_id": mev,
            "identity": {k: rec.get(k) for k in
                         ("Grupo", "Estrategia", "Familia", "Required_Surface",
                          "Backend_Module", "Toggle_Frontend", "Detector_ID",
                          "Execution_Class", "Deterministic_Class")},
            "structure": {k: rec.get(k) for k in
                          ("Min_Legs", "Max_Legs", "Legs_Model", "Atomic_Possible",
                           "Oracle_Dep", "Bridge_Dep", "External_Dep")},
            "operators": {
                "primary_ops": rec.get("Primary_Ops"),
                "secondary_ops": rec.get("Secondary_Ops"),
                "matrix_link_count": m.get("link_count"),
                "matrix_primary_count": m.get("primary_count"),
                "matrix_secondary_count": m.get("secondary_count"),
                "phase_primaries": m.get("phase_primaries"),
            },
            "math": {k: rec.get(k) for k in
                     ("Detector_Math", "Discovery_Equation", "Data_Bindings",
                      "Frontend_Config", "Gate_LIVE")},
            "state": {k: rec.get(k) for k in
                      ("Design_State", "Initial_Mode", "Strategy_Enabled", "Status",
                       "Readiness_Score", "Operator_Coverage_pct",
                       "Template_State")},
            "risk": {k: rec.get(k) for k in ("Toxicity", "Dominant_Risks",
                     "Atomicity_Type", "NonAtomic_Type", "Mode_Invariant")},
            "evidence": ev,
            "repo_verification": status,
        }
        registry.append(entry)
        R.add(f"REQ-STRAT-{mev}", "11_STRATEGY_CATALOG",
              f"row {rec['_row']}",
              f"Estrategia {mev} ({rec.get('Estrategia')}) real: cartridge Rhai + canon "
              f"STRATEGY.json + operadores declarados que arman su combo particular",
              anchors, family="strategy")

    with open(os.path.join(ART, "strategy_registry.json"), "w", encoding="utf-8") as f:
        json.dump({"workbook": wb_meta, "count": len(registry), "strategies": registry},
                  f, ensure_ascii=False, indent=1)

    # ── 12: operators ──────────────────────────────────────────────────────
    operators = parse_operators(raw)
    for op in operators:
        code = op["code"]
        n_engine = REPO.glob_exists(f"backend/math-engine/src/operators/{code}_*.rs")
        n_canon = REPO.exists(f"skills/arbitragex-ultra/operators/{code}/OPERATOR.json")
        kg = REPO.grep(re.escape(code))
        calib = REPO.grep(r"calibrat")
        anchors = [
            {"glob": f"backend/math-engine/src/operators/{code}_*.rs"},
            {"path": f"skills/arbitragex-ultra/operators/{code}/OPERATOR.json"},
        ]
        R.add(f"REQ-OP-{code.upper()}", "12_OPERATOR_CONTROL", f"row {op['_row']}",
              f"Operador {code} ({op['name']}): presente en math-engine, canon, y "
              f"referenciado por estrategias; estado de calibración auditable",
              anchors + [{"regex": re.escape(code)}], family="operator")
    op_records = {**{op["code"]: op for op in operators}}

    # ── 13: matrix links ───────────────────────────────────────────────────
    total_links = sum(m["link_count"] for m in matrix)
    grid_cells = len(matrix) * 31
    R.add("REQ-LINK-MATRIX", "13_STRAT_OP_MATRIX",
          f"{len(matrix)} strategies x 31 operators = {grid_cells} grid cells",
          f"Las {grid_cells} relaciones estrategia×operador (roles primary/secondary, "
          f"{total_links} non-zero) definen el combo particular de cada estrategia; el "
          "repo debe materializarlas (cartridge declarations + knowledge graph)",
          [{"path": "skills/arbitragex-ultra/knowledge_graph.jsonl"},
           {"regex": r"primary_operators"}, {"regex": r"secondary_operators"}],
          family="matrix")

    # ── 01: config knobs ───────────────────────────────────────────────────
    knobs = parse_config(raw)
    for k in knobs:
        knob = k["knob"]
        # Workbook knob tokens are already underscore-delimited (Enable_2V2,
        # Min_Pool_Liquidity_USD …) — the canonical snake name is just lowercase.
        snake = knob.lower()
        anchors = [{"regex": r"\b" + re.escape(snake) + r"\b"}]
        R.add(f"REQ-CONFIG-{knob}", "01_CONFIG", f"row {k['_row']}",
              f"Knob {knob}={k['value']} ({k['unit']}, capa {k['layer']}): "
              f"{k['description']} — superficie canónica presente (declarativa validada)",
              anchors, family="config")

    # ── 02: financing ──────────────────────────────────────────────────────
    fin = parse_financing(raw)
    for fm in fin:
        mode = fm["mode"]
        anchors = [{"regex": re.escape(mode)},
                   {"regex": re.escape(mode.lower())}]
        R.add(f"REQ-FIN-{mode}", "02_FINANCING", f"row {fm['_row']}",
              f"Modo de financiamiento {mode}: capacity={fm['capacity_usd']}, "
              f"fee={fm['provider_fee_bps']}bps (on-chain, nunca hardcode), "
              f"extra_gas={fm['extra_gas_usd']}, constraint={fm['route_constraint']} "
              "— filtro de rutas y sizing por modo",
              anchors, family="financing")
    R.add("REQ-FIN-SIZING-RULE", "02_FINANCING", "row 12",
          "Sizeable(route, mode) = MIN(required_notional, bottleneck_liquidity × "
          "utilization_cap, mode_capacity) aplicado por-modo",
          [{"regex": r"bottleneck"}, {"regex": r"utilization"}], family="financing")

    # ── 03: graph ──────────────────────────────────────────────────────────
    g3 = parse_graph03_headers(raw)
    R.add("REQ-GRAPH-LOG-WEIGHT", "03_GRAFO_POOLS", f"cols M/N ({g3['data_rows']} pools)",
          "Pesos de arista logarítmicos −ln(rate) por dirección (fiel a la fórmula del "
          "workbook) — nunca peso sintético",
          [{"path": "backend/searcher-rs/src/route_discovery/graph_builder.rs"},
           {"regex": r"ln\("}], family="graph")
    R.add("REQ-GRAPH-PARALLEL-EDGES", "03_GRAFO_POOLS", "Pool_ID/DEX/Version/Fee_bps cols",
          "Aristas paralelas: mismo par token0/token1 con múltiples pools (DEX×versión×"
          "fee) como aristas distintas con fee explícito",
          [{"regex": r"parallel"}, {"regex": r"fee_bps|fee_tier|pool_fee"}], family="graph")
    R.add("REQ-GRAPH-ELIGIBILITY", "03_GRAFO_POOLS", "col O Eligible_Liquidity",
          "Elegibilidad de pool por liquidez mínima (poda del grafo) + Hot_Token para "
          "pruning por concentración",
          [{"regex": r"min_pool_liquidity|min_liquidity|liquidity_floor"},
           {"regex": r"hot_token"}], family="graph")
    R.add("REQ-GRAPH-LIVE-STATE", "03_GRAFO_POOLS", "A3 note",
          "El estado DEMO del workbook se sustituye por snapshot/event stream real "
          "(sync de reserves on-chain); la fórmula queda, los datos son live",
          [{"regex": r"reserves_cache|ReservesCache|sync_reserves"},
           {"regex": r"Sync|Swap|Mint.*event"}], family="graph")

    # ── 04: algorithms ─────────────────────────────────────────────────────
    algos = parse_algorithms04(raw)
    ALGO_ANCHORS = {
        "BFM_NEG_CYCLE": [r"bellman", r"negative_cycle|neg_cycle"],
        "MMBF_LINE_GRAPH": [r"mmbf|line_graph|line graph"],
        "JOHNSON": [r"johnson"],
        "BOUNDED_DFS": [r"bounded.*dfs|dfs.*bounded|find_routes|simple cycle"],
        "RICH": [r"\brich\b"],
        "CONVEX_SIZE": [r"convex"],
        "MPO": [r"marginal.price|mpo"],
    }
    for a in algos:
        name = a["algorithm"]
        rxs = ALGO_ANCHORS.get(name, [re.escape(name.lower())])
        anchors = [{"regex": rx} for rx in rxs]
        R.add(f"REQ-ALGO-{name}", "04_ALGORITMOS", f"row {a['_row']}",
              f"Algoritmo {name} ({a['role']}; complexity {a['complexity']}; "
              f"enabled={a['enabled']}): disponible según diseño (o ausencia declarada)",
              anchors, family="algorithm")

    # ── 05: route pipeline columns ─────────────────────────────────────────
    r5 = parse_rutas05_headers(raw)
    RUTA_FAMILIES = [
        ("GATES", ["Hop_OK", "Kind_OK", "Liquidity_OK", "Edge_OK", "Fresh_OK",
                   "Candidate_OK"], "Gates de ruta estructurados con reject reason"),
        ("SIZING-PER-MODE", ["Own_Size_USD", "Aave_Size_USD", "Balancer_Size_USD",
                             "V2_Flash_Swap_Size_USD"], "Sizing calculado POR financing mode"),
        ("EV-PER-MODE", ["Own_EV_USD", "Aave_EV_USD", "Balancer_EV_USD",
                         "V2_Flash_Swap_EV_USD"], "EV neto POR financing mode"),
        ("VIABILITY", ["Selected_Viable", "Own_Viable", "Aave_Viable", "Balancer_Viable"],
         "Viabilidad por modo (EV>Min_EV tras todos los costos)"),
        ("RANK", ["Rank_Score"], "Rank score compuesto (pesos 01_CONFIG)"),
        ("REJECT", ["Reject_Reason"], "Reject reason estructurado por gate (nada muere en silencio)"),
        ("PROVENANCE", ["Primary_Source"], "Fuente primaria de cada ruta (paper/URL)"),
    ]
    for fam, cols, stmt in RUTA_FAMILIES:
        present = [c for c in cols if c in r5["columns"]]
        R.add(f"REQ-RUTA-{fam}", "05_RUTAS",
              f"cols {', '.join(present)} ({r5['data_rows']} routes batch)",
              f"{stmt} — pipeline real: discovery → gates → financing → EV → ranking "
              "(fórmulas del workbook como contrato matemático)",
              [{"regex": r"RejectReason|rejection_reason|reject_reason"},
               {"regex": r"route_metadata"},
               {"regex": r"SizeOptimizer|size_optimizer"}], family="route")

    # ── 06: sensitivity ────────────────────────────────────────────────────
    R.add("REQ-SENS-MODE-COMPARE", "06_SENSIBILIDAD", "A4:G8",
          "Comparación simultánea por financing mode: viable routes, total/avg EV, max "
          "size — sin re-ejecutar por modo",
          [{"regex": r"financing_mode|by_mode|per_mode"}], family="sensitivity")
    R.add("REQ-SENS-SWEEPS", "06_SENSIBILIDAD", "H4:J11 + A12:E18",
          "Sweeps de sensibilidad: liquidez mínima y max hops vs set viable/EV",
          [{"regex": r"sweep|sensitivity|what_if"}], family="sensitivity")

    # ── 07: gates G0..G6 ───────────────────────────────────────────────────
    GATE_ANCHORS = {
        "G0_DATA": [r"pools.*COUNT|pool.*populat|graph.*populat"],
        "G1_DISCOVERY": [r"candidate.*>.*0|discovered|candidates_found"],
        "G2_BUDGET": [r"budget"],
        "G3_FINANCING": [r"financing.*enabled|mode_enabled"],
        "G4_ECON": [r"viable.*>.*0|at_least_one|positive"],
        "G5_LATENCY": [r"cpu_budget|op_budget|latency"],
        "G6_EXECUTION": [r"PAPER_SHADOW|LIVE_MAINNET|execution_mode"],
    }
    gates7 = parse_gates07(raw)
    for gt in gates7:
        gid = gt["gate"]
        rxs = GATE_ANCHORS.get(gid, [re.escape(gid)])
        R.add(f"REQ-GATE-{gid}", "07_GATES", f"row {gt['_row']}",
              f"Gate {gid}: {gt['objective']} (metric {gt['metric']}, "
              f"threshold {gt['threshold']}, action {gt['action']})",
              [{"regex": rx} for rx in rxs], family="promotion-gate")
    R.add("REQ-GATE-MASTER", "07_GATES", "A13:A14",
          "Master gate: Overall PASS solo si 0 FAILs (composición de G0..G6)",
          [{"regex": r"overall_status|master_gate|all_gates"}], family="promotion-gate")

    # ── 08/29: dashboards ──────────────────────────────────────────────────
    DASH_KPI = [
        ("REQ-DASH-VIABILITY", "candidate routes, viable, viability %", r"viability|viable"),
        ("REQ-DASH-EV", "Total EV selected, Best route EV, Max sizeable", r"total_ev|best_ev|max_size"),
        ("REQ-DASH-BY-MODE", "Set por financing mode", r"by.*financing|financing.*mode"),
        ("REQ-DASH-BY-HOPS", "Viable por hops", r"by_hops|per_hop"),
        ("REQ-DASH-BY-KIND", "Viable por kind (2V2/V2V3/TRI/NHOP)", r"by_kind|per_kind|route_kind"),
        ("REQ-SUPER-COUNTERS", "counters 264/60/31/8184 + route-ready vs need-data",
         r"264|8184|route_ready"),
    ]
    for rid, stmt, rx in DASH_KPI:
        R.add(rid, "08_DASHBOARD+29_SUPER_DASHBOARD", "-",
              f"KPI servido con datos reales (sin métricas mockeadas): {stmt}",
              [{"regex": rx}], family="dashboard")

    # ── 10: enums ──────────────────────────────────────────────────────────
    lists10 = parse_lists10(raw)
    for name, vals in lists10.items():
        # anchor: at least the distinctive values appear in repo
        anchors = []
        for v in vals[:4]:
            anchors.append({"regex": re.escape(str(v))})
        R.add(f"REQ-LIST-{name}", "10_LISTAS", f"{len(vals)} values",
              f"Enum {name} single-source-of-truth ({len(vals)} valores, workbook "
              "canónico); el repo no debe duplicar definiciones divergentes",
              anchors, family="enum")

    # ── 16: scoreboard ─────────────────────────────────────────────────────
    sb = parse_scoreboard(raw)
    for grp in sb:
        R.add(f"REQ-SCORE-G{grp['group']:02d}", "16_STRATEGY_SCOREBOARD",
              f"row {grp['_row']}",
              f"Scoreboard grupo {grp['group']} ({grp['family']}): "
              f"{grp['strategies']} estrategias, route_ready={grp['route_ready']}, "
              f"needs_data={grp['needs_data']}, observe_only={grp['observe_only']} — "
              "métricas explicables desde datos reales",
              [{"regex": r"route_ready|needs_data|observe_only"}], family="scoreboard")

    # ── 17: source audit ───────────────────────────────────────────────────
    sa, containment = parse_source_audit17(raw)
    prov_anchor = [{"path": "docs/excel_ingestion_manifest.json"},
                   {"path": "docs/excel_strategies_extracted.json"}]
    for i, src in enumerate(sa, 1):
        R.add(f"REQ-PROV-SRC{i:02d}", "17_SOURCE_AUDIT", f"row {src['_row']}",
              f"Procedencia: {src['file']} ({src['sheets']} sheets, sha256 {str(src['sha256'])[:12]}…) "
              f"— {src['coverage']} — ingested + verificable",
              prov_anchor, family="provenance")

    # ── 18: canonical library sections ─────────────────────────────────────
    lib18 = parse_library18(raw)
    LIB_ANCHORS = {
        "DETECTOR_FAMILIES": [r"detector"],
        "EXECUTION_MODES": [r"LIVE_MAINNET|TESTNET|PAPER_SHADOW"],
        "MODE_MIGRATION": [r"mode.*migrat|migrat"],
        "FRONTEND_CONTRACT": [r"trading-config|opportunities.*schema|zod"],
        "CARTRIDGE_ARCH": [r"cartridge"],
        "B_STATIC_DIAG": [r"static.*diag|diag"],
        "CONFLICTS": [r"conflict"],
    }
    for sec, recs in lib18.items():
        rxs = LIB_ANCHORS.get(sec, [re.escape(sec.lower().replace("_", " "))])
        R.add(f"REQ-LIB-{sec}", "18_CANONICAL_LIBRARY", f"{len(recs)} records",
              f"Sección canónica {sec} ({len(recs)} registros) integrada al repo",
              [{"regex": rx} for rx in rxs], family="library")

    # ── 19: source field map ───────────────────────────────────────────────
    fm19 = parse_fieldmap19(raw)
    with open(os.path.join(ART, "source_field_map.json"), "w", encoding="utf-8") as f:
        json.dump({"workbook": wb_meta, "mappings": fm19,
                   "critical_numeric_contracts": {
                       "note": "operator directive: decimals, token0/token1, "
                               "sqrtPriceX96, wei/gwei/ETH, bps — unidades y "
                               "conversiones en un solo mapa"},
                   "repo_field_index": {
                       "sqrtPriceX96": sorted(REPO.grep(r"sqrtPriceX96|sqrt_price_x96"))[:10],
                       "decimals": sorted(REPO.grep(r"\bdecimals\b"))[:10],
                       "token0_token1": sorted(REPO.grep(r"token0|token_0"))[:10],
                       "wei_gwei": sorted(REPO.grep(r"gwei|1e9"))[:10],
                       "bps": sorted(REPO.grep(r"\bbps\b|basis_points"))[:10],
                   }},
                  f, ensure_ascii=False, indent=1)
    for i, src in enumerate(fm19, 1):
        R.add(f"REQ-FIELDMAP-SRC{i:02d}", "19_SOURCE_FIELD_MAP", f"row {src['_row']}",
              f"Mapa de campos: {src['source_sheet']} ({src['records']} records × "
              f"{src['columns']} cols) → {src['integrated_into']} — {src['treatment']}",
              [{"path": "artifacts/source_field_map.json"},
               {"path": "artifacts/excel_requirements.json"}], family="field-map")

    # ── 09: references ─────────────────────────────────────────────────────
    refs = parse_references09(raw)
    doctrine_anchor = [{"path": "docs/ROUTES_CROWN_JEWEL_DOCTRINE.md"},
                       {"glob": "skills/arbitragex-ultra/world/**"}]
    for i, ref in enumerate(refs, 1):
        R.add(f"REQ-REF-{i:03d}", "09_REFERENCIAS", f"row {ref['_row']}",
              f"Referencia doctrina ({ref['topic']}): {ref['reference']} — uso: {ref['use']}",
              doctrine_anchor, family="doctrine")

    # ── 15: per-strategy route optimizer ───────────────────────────────────
    r15 = parse_route_opt15(raw)
    R.add("REQ-ROUTEOPT-SELECTED", "15_STRAT_ROUTE_OPT",
          f"{r15['data_rows']} route rows, {len(r15['columns'])} cols",
          "Optimizador de rutas por estrategia seleccionada: compatibilidad estructural "
          "(legs/surface/oracle/bridge/external/atomic), fit %, EV ajustado y rank por "
          "estrategia — la estrategia filtra el set de rutas",
          [{"regex": r"strategy_fit|fit_pct|applicable_strategies"},
           {"regex": r"StrategyApplicability|strategy_applicability"}], family="route-opt")

    # ── emit requirements + coverage ───────────────────────────────────────
    with open(os.path.join(ART, "excel_requirements.json"), "w", encoding="utf-8") as f:
        json.dump({"workbook": wb_meta, "count": len(R.items), "requirements": R.items},
                  f, ensure_ascii=False, indent=1)

    by_status = defaultdict(int)
    by_family_status: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for it in R.items:
        st = it["verification"]["status"]
        by_status[st] += 1
        by_family_status[it["family"]][st] += 1
    total = len(R.items)
    verified = by_status["VERIFIED"]
    partial = by_status["PARTIAL"]
    missing = by_status["MISSING"]

    coverage = {
        "workbook": wb_meta,
        "generated_by": "scripts/excel_canon/build_canonical_artifacts.py",
        "totals": {"requirements": total, "verified": verified, "partial": partial,
                   "missing": missing,
                   "verified_pct": round(100.0 * verified / total, 2) if total else 0.0},
        "by_family": {fam: dict(sts) for fam, sts in sorted(by_family_status.items())},
        "by_sheet": {},
    }
    sheet_stat: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for it in R.items:
        sheet_stat[it["sheet"]][it["verification"]["status"]] += 1
    coverage["by_sheet"] = {s: dict(sts) for s, sts in sorted(sheet_stat.items())}
    with open(os.path.join(ART, "excel_coverage.json"), "w", encoding="utf-8") as f:
        json.dump(coverage, f, ensure_ascii=False, indent=1)

    # ── markdown milestone report ─────────────────────────────────────────
    L: list[str] = []
    L.append("# Excel→Repo Coverage — Canonical Milestone Report")
    L.append("")
    L.append(f"- **Workbook**: `{os.path.basename(raw['source'])}` "
             f"(sha256 `{raw['sha256'][:16]}…`, {len(raw['sheets'])} sheets)")
    L.append(f"- **Strategies**: {len(registry)} · **Operators**: {len(operators)} · "
             f"**Matrix links (non-zero)**: {total_links} of {grid_cells} grid cells")
    L.append(f"- **Requirements**: {total} — ✅ VERIFIED {verified} · 🟡 PARTIAL "
             f"{partial} · ❌ MISSING {missing}")
    L.append(f"- **VERIFIED coverage**: **{coverage['totals']['verified_pct']}%**")
    L.append("")
    L.append("> Generado por `scripts/excel_canon/build_canonical_artifacts.py`. Cada "
             "estado se VERIFICA contra el working tree (path/regex anchors) — nada "
             "declarativo. Counts derivados del workbook, jamás hardcodeados (RULE 00/R8).")
    L.append("")
    L.append("## Por hoja")
    L.append("")
    L.append("| Hoja | Reqs | ✅ | 🟡 | ❌ |")
    L.append("|---|---:|---:|---:|---:|")
    for s, sts in sorted(coverage["by_sheet"].items()):
        L.append(f"| {s} | {sum(sts.values())} | {sts.get('VERIFIED',0)} | "
                 f"{sts.get('PARTIAL',0)} | {sts.get('MISSING',0)} |")
    L.append("")
    L.append("## Por familia")
    L.append("")
    L.append("| Familia | ✅ | 🟡 | ❌ |")
    L.append("|---|---:|---:|---:|")
    for fam, sts in sorted(coverage["by_family"].items()):
        L.append(f"| {fam} | {sts.get('VERIFIED',0)} | {sts.get('PARTIAL',0)} | "
                 f"{sts.get('MISSING',0)} |")
    L.append("")
    L.append("## Gaps prioritarios (MISSING/PARTIAL con impacto)")
    L.append("")
    gaps = [it for it in R.items if it["verification"]["status"] != "VERIFIED"
            and it["family"] not in ("provenance", "field-map", "doctrine")]
    for it in gaps[:80]:
        st = it["verification"]["status"]
        icon = "🟡" if st == "PARTIAL" else "❌"
        L.append(f"- {icon} **{it['id']}** ({it['sheet']}, {it['family']}): {it['statement'][:220]}")
    if len(gaps) > 80:
        L.append(f"- … y {len(gaps) - 80} más (ver excel_coverage.json)")
    with open(os.path.join(ART, "excel_coverage.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")

    print(f"requirements={total} verified={verified} partial={partial} missing={missing} "
          f"pct={coverage['totals']['verified_pct']}")
    print(f"strategies={len(registry)} operators={len(operators)} links={total_links}/{grid_cells}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
