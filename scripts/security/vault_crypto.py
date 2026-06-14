#!/usr/bin/env python3
import os
import sys
import json
import hashlib
import argparse
import base64
from pathlib import Path
try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
except ImportError:
    print("Error: cryptography package is required. Run 'pip install cryptography'", file=sys.stderr)
    sys.exit(1)
try:
    import openpyxl
except ImportError:
    print("Error: openpyxl package is required. Run 'pip install openpyxl'", file=sys.stderr)
    sys.exit(1)

def extract_master_key_from_excel(excel_path: str) -> bytes:
    """Reads the Excel file and generates a stable 32-byte AES key based on all secrets."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Collect all secrets in a stable, deterministic order
    secrets_array = []
    for sheet_name in sorted(wb.sheetnames):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row[0] is not None and row[1] is not None:
                key = str(row[0]).strip()
                val = str(row[1]).strip()
                
                # Skip header rows
                if key.lower() in ('variable', 'header', 'comando', 'key', 'nombre'):
                    continue
                if val.lower() in ('valor', 'value', 'curl', 'notas'):
                    continue
                
                secrets_array.append(f"{key}={val}")
    
    if not secrets_array:
        raise ValueError("No secrets found in the Excel file")
        
    # Create a stable string representation
    stable_string = "\n".join(secrets_array)
    
    # Use SHA-256 to derive a 32-byte key for AES-256
    # We use a salt to ensure this key is specific to ArbitrageX
    salt = b"ArbitrageX_v2_Master_Key_Salt_2026"
    
    # PBKDF2 or similar would be better, but a double SHA-256 is sufficient here
    # since the input is already high-entropy (contains 60+ strong passwords)
    h = hashlib.sha256()
    h.update(salt)
    h.update(stable_string.encode('utf-8'))
    
    return h.digest()

def encrypt_file(input_path: str, output_path: str, key: bytes):
    """Encrypts a file using AES-256-GCM."""
    with open(input_path, 'rb') as f:
        plaintext = f.read()
        
    aesgcm = AESGCM(key)
    nonce = os.urandom(12) # 96-bit nonce is standard for GCM
    
    # Encrypt (GCM appends the 16-byte auth tag to the ciphertext automatically)
    ciphertext = aesgcm.encrypt(nonce, plaintext, None)
    
    # We store the nonce + ciphertext
    with open(output_path, 'wb') as f:
        f.write(nonce + ciphertext)
        
    print(f"Successfully encrypted {input_path} -> {output_path}")

def decrypt_file(input_path: str, output_path: str, key: bytes):
    """Decrypts a file using AES-256-GCM."""
    with open(input_path, 'rb') as f:
        data = f.read()
        
    if len(data) < 28: # 12 bytes nonce + 16 bytes tag + min 0 bytes ciphertext
        raise ValueError("Encrypted file is too small or corrupted")
        
    nonce = data[:12]
    ciphertext = data[12:]
    
    aesgcm = AESGCM(key)
    
    try:
        plaintext = aesgcm.decrypt(nonce, ciphertext, None)
    except Exception as e:
        raise ValueError(f"Decryption failed (wrong key or corrupted file): {e}")
        
    # Ensure directory exists
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    
    with open(output_path, 'wb') as f:
        f.write(plaintext)
        
    # Make sure the decrypted file is readable only by the owner
    os.chmod(output_path, 0o600)

def main():
    parser = argparse.ArgumentParser(description="ArbitrageX v2 Vault Crypto Tool")
    parser.add_argument("action", choices=["encrypt", "decrypt", "get-key-hash"], help="Action to perform")
    parser.add_argument("--excel", required=True, help="Path to ArbitrageX_Secrets_Config.xlsx")
    parser.add_argument("--in-file", help="Input file (e.g., .env or .env.enc)")
    parser.add_argument("--out-file", help="Output file")
    
    args = parser.parse_args()
    
    try:
        key = extract_master_key_from_excel(args.excel)
        
        if args.action == "get-key-hash":
            # Just print a hash of the key to verify it matches without exposing it
            key_hash = hashlib.sha256(key).hexdigest()[:16]
            print(f"Master Key Hash: {key_hash}")
            sys.exit(0)
            
        if not args.in_file or not args.out_file:
            print("Error: --in-file and --out-file are required for encrypt/decrypt", file=sys.stderr)
            sys.exit(1)
            
        if args.action == "encrypt":
            encrypt_file(args.in_file, args.out_file, key)
        elif args.action == "decrypt":
            decrypt_file(args.in_file, args.out_file, key)
            
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
