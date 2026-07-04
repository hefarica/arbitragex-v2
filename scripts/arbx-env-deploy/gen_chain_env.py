#!/usr/bin/env python3
"""
Generate the 5 chain artifacts from the Excel "Chain Builder" sheet.

Reads:  Chain Builder col B (v ACTIVE) + the 3 source sheets (RPC Providers,
        _RED_lookup) via the canonical chain_catalog.json (or directly).
Emits (to <out> dir, gitignored-eligible fragments for review):
  1. rpc_chains_fragment.env      — RPC_HTTP_<id> + RPC_WS_<id> (multi-provider, SHUFFLED)
  2. enabled_chains_fragment.env  — ARBX_ENABLED_CHAINS=<comma list of active chain_ids>
  3. chains_seed.sql              — INSERT INTO chains (chain_id, name, native_currency, ...)
  4. factories_seed.sql           — INSERT INTO factories (dex_id, chain_id, address)
                                    [dex_id resolved via subquery to dexes.name — FK-safe]
  5. tokens_reference.md          — top tokens per chain (WNative + USDC-equiv)

Direction: Excel -> fragments (read-only on workbook). Applying fragments to
.env Production / VPS / PG migrations is a separate operator-gated step
(RunFullSyncCycle + psql).

Privacy: multi-provider CSV + provider ORDER IS SHUFFLED per run, so the
first provider in the value does not fingerprint the operator's infra.

Read-only on the workbook. No secret is printed to stdout (only lengths).
paper_mode is never touched (it lives in .env Production, not in fragments).
"""
import argparse
import json
import random
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# Canonical CHAIN_IDS (mirrors gen_rpc_env_from_xlsx.py + seed 043) — FIX GAP 1.
CHAIN_IDS = {
    "Ethereum Mainnet": 1, "Optimism": 10, "BSC Mainnet": 56, "Gnosis": 100,
    "Polygon Mainnet": 137, "Base": 8453, "Arbitrum One": 42161,
    "Avalanche": 43114, "Linea": 59144, "Scroll": 534352, "Blast": 81457,
}
CHAIN_META = {
    1:    ("ethereum", "ETH",   "https://etherscan.io",            12000),
    10:   ("optimism", "ETH",   "https://optimistic.etherscan.io",  2000),
    56:   ("bsc",      "BNB",   "https://bscscan.com",              3000),
    100:  ("gnosis",   "xDAI",  "https://gnosisscan.io",            5000),
    137:  ("polygon",  "MATIC", "https://polygonscan.com",          2000),
    8453: ("base",     "ETH",   "https://basescan.org",             2000),
    42161:("arbitrum", "ETH",   "https://arbiscan.io",               250),
    43114:("avalanche","AVAX",  "https://snowtrace.io",             2000),
    59144:("linea",    "ETH",   "https://lineascan.build",         12000),
    534352:("scroll",  "ETH",   "https://scrollscan.com",           3000),
    81457:("blast",    "ETH",   "https://blastscan.io",             2000),
}
# Factories per chain (dex_name -> router address) — canonical mainnet values.
# FIX GAP 2: factories_seed.sql resolves dex_name -> dex_id UUID via subquery
# (SELECT id FROM dexes WHERE name=...) so the FK is always satisfied.
FACTORIES = {
    1:    {"UniswapV2": "0x5C69bEe701ef814a2B6a3EDD4B1652CB9cc5aA6f",
           "UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
           "SushiSwap": "0xC0AEe478e3658e2610c5F7A4A2E1777cE9e4f2Ac"},
    10:   {"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
           "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
    56:   {"PancakeSwap V2": "0x1097053Fd5911a4863cA7D0e6F3C73a8B2CDA8b9",
           "PancakeSwap V3": "0x0BFbCF9fa4f9C56B0F40a671Ad90E3DC94D20d4e",
           "BiSwap": "0x3a6d8cA21D1CF76F653A67577FA0FB271661792C"},
    137:  {"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
           "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
    8453: {"UniswapV3": "0x33128a8fC17869897dcEA68d25cD9Ec44D11BbfA",
           "Aerodrome": "0x33360F37492Ea44090b89FF2cFF92Bc399938E1"},
    42161:{"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
           "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
}
TOKENS_REF = {  # WNative + USDC-equivalent per chain
    1:    ("WETH", "0xC02aaA39b223FE8D0A0e5C4F27eAD9083C756Cc2", "USDC", "0xA0b86991c6218b36c1d19D4a2e9Eb0cE3606eB48"),
    10:   ("WETH", "0x4200000000000000000000000000000000000006", "USDC", "0x7F5c764cBc14f9669B88837ca1490cCa17c31607"),
    137:  ("WMATIC","0x0d500B1d8E8eF31E21C99d1Db9A6444d3ADf1270", "USDC", "0x2791Bca1f2de4661ED88A30C99A7a9449Aa84174"),
    8453: ("WETH", "0x4200000000000000000000000000000000000006", "USDC", "0x833589fCD6eDb6E08f4c7C32D4f71b54bdA20D63"),
    42161:("WETH", "0x82aF49447D8a07e3bd95BD0d56f35241523fBab1", "USDC", "0xaf88d065e77c8cC2239327C5EDb3A432268e5831"),
    56:   ("WBNB", "0xbb4CdB9CBd36B01bD1cBaEBF2De08d9173bc095c", "BUSD", "0xe9e7CEA3DedC5394299B3f8f10F1Bb15Bb1b7b15"),
}


def load_catalog(wb):
    """Read RPC Providers -> chain->{HTTP,WSS}->{provider:url}."""
    ws = wb["RPC Providers"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chain, proto, prov, url = (list(row) + [None] * 4)[:4]
        if not all([chain, proto, prov, url]):
            continue
        out.setdefault(str(chain).strip(), {"HTTP": {}, "WSS": {}})
        out[str(chain).strip()][str(proto).strip()][str(prov).strip()] = str(url).strip()
    return out


def shuffle_csv(providers, seed=None):
    """Multi-provider CSV with SHUFFLED order (privacy: no fixed first provider)."""
    items = list(providers.items())
    rng = random.Random(seed)
    rng.shuffle(items)
    return ",".join(f"{prov}={url}" for prov, url in items)


def main():
    # Windows cp1252 can't encode some Unicode (arrows, checkmarks) — force utf-8 so
    # informational prints don't crash on the final summary line.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default="C:/Users/HFRC/Downloads/ArbitrageX_Unified_Config.xlsm")
    ap.add_argument("--out", default=".", help="dir for the 5 fragments")
    ap.add_argument("--seed", type=int, default=None, help="shuffle seed (None=random per run)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True, keep_vba=True)
    if "Chain Builder" not in wb.sheetnames:
        sys.exit("ERROR: 'Chain Builder' sheet not found in workbook")
    ws = wb["Chain Builder"]

    # Read active chains (col B == "✓" — the checkmark the operator marks)
    # Accept BOTH "✓" (Unicode, what the data validation produces) AND "v" (ASCII fallback)
    active = []  # [(chain_name, chain_id)]
    for row in ws.iter_rows(min_row=4, values_only=True):
        chain, mark, cid = (list(row) + [None] * 3)[:3]
        mark_s = str(mark or "").strip()
        if mark_s in ("✓", "v", "V", "x", "X") and chain and cid:
            active.append((str(chain).strip(), int(cid)))
    if not active:
        sys.exit("No chains marked ACTIVE in Chain Builder col B. Mark a checkmark in col B then re-run.")

    rpcs = load_catalog(wb)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    rpc_lines, chains_sql, factories_sql, tokens_md = [], [], [], []
    enabled_ids = [str(cid) for _, cid in active]

    for chain, cid in active:
        rpc = rpcs.get(chain)
        if not rpc:
            print(f"WARN: {chain} ({cid}) not in RPC Providers — skipped", file=sys.stderr)
            continue
        http, wss = rpc.get("HTTP", {}), rpc.get("WSS", {})
        if not http:
            print(f"WARN: {chain} ({cid}) has no HTTP RPCs — skipped", file=sys.stderr)
            continue
        rpc_lines.append(f"RPC_HTTP_{cid}={shuffle_csv(http, args.seed)}")
        if wss:
            rpc_lines.append(f"RPC_WS_{cid}={shuffle_csv(wss, args.seed)}")

        # chains.sql — uses CHAIN_META (canonical name/native/explorer) — FIX GAP 1
        name, native, explorer, _ = CHAIN_META.get(cid, (chain.lower(), "?", "?", 0))
        chains_sql.append(
            f"  ({cid}, '{name}', '{native}', '{explorer}', true)"
        )

        # factories.sql — FK-safe via subquery to dexes.name — FIX GAP 2
        if cid in FACTORIES:
            for dex_name, addr in FACTORIES[cid].items():
                # Resolve dex_name -> dex_id UUID at apply-time (subquery), not here.
                # This makes the SQL independent of the dexes.id UUID value.
                factories_sql.append(
                    f"  ((SELECT id FROM dexes WHERE name='{dex_name}'), {cid}, '{addr}')"
                )

        # tokens reference
        if cid in TOKENS_REF:
            sym1, addr1, sym2, addr2 = TOKENS_REF[cid]
            tokens_md.append(f"| {cid} | {chain} | {sym1} `{addr1}` | {sym2} `{addr2}` |")

    # Write the 5 fragments
    (out / "rpc_chains_fragment.env").write_text("\n".join(rpc_lines) + "\n")
    (out / "enabled_chains_fragment.env").write_text(f"ARBX_ENABLED_CHAINS={','.join(enabled_ids)}\n")
    (out / "chains_seed.sql").write_text(
        "-- Generated by gen_chain_env.py — review before applying.\n"
        "-- Idempotent: ON CONFLICT DO NOTHING.\n"
        "INSERT INTO chains (chain_id, name, native_currency, explorer_url, is_active) VALUES\n"
        + ",\n".join(chains_sql)
        + "\nON CONFLICT (chain_id) DO NOTHING;\n"
    )
    (out / "factories_seed.sql").write_text(
        "-- Generated by gen_chain_env.py.\n"
        "-- dex_id resolved via subquery (FK-safe, UUID-independent).\n"
        "INSERT INTO factories (dex_id, chain_id, address) VALUES\n"
        + ",\n".join(factories_sql)
        + "\nON CONFLICT (chain_id, address) DO NOTHING;\n"
    ) if factories_sql else (out / "factories_seed.sql").write_text("-- No factories for the active chains.\n")
    (out / "tokens_reference.md").write_text(
        "# Tokens reference per active chain\n\n"
        "| chain_id | chain | WNative | USDC-equiv |\n|---|---|---|---|\n"
        + "\n".join(tokens_md) + "\n"
    )

    print(f"OK — {len(active)} chain(s) active: {enabled_ids}")
    print(f"fragments in {out.resolve()}:")
    for f in ["rpc_chains_fragment.env", "enabled_chains_fragment.env",
              "chains_seed.sql", "factories_seed.sql", "tokens_reference.md"]:
        p = out / f
        print(f"  {f}: {p.stat().st_size} bytes")
    print("\nNext (operator-gated): review fragments -> apply via RunFullSyncCycle (.env) + psql (SQL).")


if __name__ == "__main__":
    main()
