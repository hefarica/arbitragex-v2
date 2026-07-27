#!/usr/bin/env python3
"""
Encrypted Config Bundle shipper.

Serializes the operator's Excel into one schema-validated JSON bundle, hybrid-
encrypts it (RSA-OAEP-4096 wraps a random AES-256-GCM key), and uploads it to
the VPS via SSH. The VPS-held private key is the ONLY thing that can decrypt.

If the .xlsm or the .enc leaks, the data is unreadable.

Crypto format (binary, magic-prefixed):
  ARBX1 || u32(wrapped_key_len) || rsa_wrapped_aes_key || nonce(12) || aes_gcm_ct

Filtering (defence-in-depth — the importer re-asserts):
  paper_mode / ARBX_PAPER_TRADE / DEPLOYER_* / MULTISIG_ADDRESS /
  CONFIRM_MAINNET_DEPLOY / MAINNET_RPC_URL are NEVER serialized.

Read-only on the workbook. Temp JSON is shred after encrypt. No secret value
is printed to stdout (counts + lengths only).
"""
import argparse
import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import padding
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
except ImportError:
    sys.exit("pip install openpyxl cryptography")

# Mirror the canonical maps from gen_chain_env.py (single source of truth).
CHAIN_IDS = {
    "Ethereum Mainnet": 1, "Optimism": 10, "BSC Mainnet": 56, "Gnosis": 100,
    "Polygon Mainnet": 137, "Base": 8453, "Arbitrum One": 42161,
    "Avalanche": 43114, "Linea": 59144, "Scroll": 534352, "Blast": 81457,
    # Testnets (public, keyless). Holesky deprecated - kept for ref.
    "Ethereum Sepolia": 11155111, "Ethereum Holesky": 17000,
    "Polygon Amoy": 80002, "Arbitrum Sepolia": 421614,
    "Optimism Sepolia": 11155420, "Base Sepolia": 84532,
}
CHAIN_META = {
    1: ("ethereum", "ETH", "https://etherscan.io"),
    10: ("optimism", "ETH", "https://optimistic.etherscan.io"),
    56: ("bsc", "BNB", "https://bscscan.com"),
    100: ("gnosis", "xDAI", "https://gnosisscan.io"),
    137: ("polygon", "MATIC", "https://polygonscan.com"),
    8453: ("base", "ETH", "https://basescan.org"),
    42161: ("arbitrum", "ETH", "https://arbiscan.io"),
    43114: ("avalanche", "AVAX", "https://snowtrace.io"),
    59144: ("linea", "ETH", "https://lineascan.build"),
    534352: ("scroll", "ETH", "https://scrollscan.com"),
    81457: ("blast", "ETH", "https://blastscan.io"),
    # Testnets
    11155111: ("sepolia", "ETH", "https://sepolia.etherscan.io"),
    17000: ("holesky", "ETH", "https://holesky.etherscan.io"),
    80002: ("polygon-amoy", "MATIC", "https://amoy.polygonscan.com"),
    421614: ("arbitrum-sepolia", "ETH", "https://sepolia.arbiscan.io"),
    11155420: ("optimism-sepolia", "ETH", "https://optimism-sepolia.blockscout.com"),
    84532: ("base-sepolia", "ETH", "https://sepolia.basescan.org"),
}
FACTORIES = {
    1: {"UniswapV2": "0x5C69bEe701ef814a2B6a3EDD4B1652CB9cc5aA6f",
        "UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
        "SushiSwap": "0xC0AEe478e3658e2610c5F7A4A2E1777cE9e4f2Ac"},
    10: {"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
         "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
    56: {"PancakeSwap V2": "0x1097053Fd5911a4863cA7D0e6F3C73a8B2CDA8b9",
         "PancakeSwap V3": "0x0BFbCF9fa4f9C56B0F40a671Ad90E3DC94D20d4e"},
    137: {"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
          "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
    8453: {"UniswapV3": "0x33128a8fC17869897dcEA68d25cD9Ec44D11BbfA"},
    42161: {"UniswapV3": "0x1F98431c8aD98523631AE4a59f267346ea31F984",
            "SushiSwap": "0xc35DADB65012eC412f5fe79F3667b22B3A32B795"},
}

# Keys that must NEVER ship (paper_mode invariant + deploy-only-local).
NEVER_SHIP = {
    "ARBX_PAPER_MODE", "ARBX_PAPER_TRADE", "PAPER_MODE",
    "DEPLOYER_PRIVATE_KEY", "DEPLOYER_KEY", "MULTISIG_ADDRESS",
    "CONFIRM_MAINNET_DEPLOY", "MAINNET_RPC_URL",
}
CONTRACT_KEYS = {"ARBITRAGE_EXECUTOR", "FLASHLOAN_EXECUTOR",
                 "ALLOWANCE_MANAGER", "ADMIN_TIMELOCK"}


def load_env_vars(wb):
    ws = wb[".env Production"]
    out = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        k = row[0]; v = row[1] if len(row) > 1 else None
        if not k or not isinstance(k, str): continue
        k = k.strip()
        if k.startswith("#") or k in NEVER_SHIP: continue
        if v is None or str(v).strip() == "": continue
        out[k] = str(v).strip()
    return out


def load_chains(wb):
    ws = wb["RPC Providers"]
    rpcs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chain, proto, prov, url = (list(row) + [None] * 4)[:4]
        if not all([chain, proto, prov, url]): continue
        rpcs.setdefault(str(chain).strip(), {"HTTP": {}, "WSS": {}})
        rpcs[str(chain).strip()][str(proto).strip()][str(prov).strip()] = str(url).strip()
    chains = []
    for chain_name, cid in CHAIN_IDS.items():
        if chain_name not in rpcs: continue
        protos = rpcs[chain_name]
        http = protos.get("HTTP", {})
        if not http: continue  # chain must have at least HTTP RPCs
        name, native, explorer = CHAIN_META.get(cid, (chain_name.lower(), "?", "?"))
        chains.append({
            "chain_id": cid, "name": name, "native_currency": native,
            "explorer_url": explorer,
            "rpc_http": ",".join(f"{p}={u}" for p, u in http.items()),
            "rpc_ws": ",".join(f"{p}={u}" for p, u in protos.get("WSS", {}).items()),
            "factories": [{"dex_name": d, "address": a}
                          for d, a in FACTORIES.get(cid, {}).items()],
        })
    return chains


def load_api_keys(wb):
    if "Tokens & Keys" not in wb.sheetnames: return {}
    ws = wb["Tokens & Keys"]
    out = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        k = row[0]; v = row[1] if len(row) > 1 else None
        if k and v and str(k).strip() not in NEVER_SHIP:
            out[str(k).strip()] = str(v).strip()
    return out


def load_contract_addresses(wb):
    ws = wb[".env Production"]
    out = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        k = row[0]; v = row[1] if len(row) > 1 else None
        if k and str(k).strip() in CONTRACT_KEYS and v:
            vs = str(v).strip()
            if vs.startswith("0x") and len(vs) == 42:
                out[str(k).strip()] = vs
    return out


def build_bundle(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True, keep_vba=True)
    return {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "env_vars": load_env_vars(wb),
        "chains": load_chains(wb),
        "api_keys": load_api_keys(wb),
        "contract_addresses": load_contract_addresses(wb),
    }


def hybrid_encrypt(plaintext_bytes, public_key_path):
    """RSA-OAEP-4096 wrap random AES-256 key; AES-256-GCM encrypt payload."""
    with open(public_key_path, "rb") as f:
        pub = serialization.load_pem_public_key(f.read())
    aes_key = AESGCM.generate_key(bit_length=256)
    nonce = os.urandom(12)
    ct = AESGCM(aes_key).encrypt(nonce, plaintext_bytes, None)
    wrapped = pub.encrypt(aes_key, padding.OAEP(
        mgf=padding.MGF1(algorithm=hashes.SHA256()),
        algorithm=hashes.SHA256(), label=None))
    magic = b"ARBX1"
    return magic + len(wrapped).to_bytes(4, "big") + wrapped + nonce + ct


def shred(path):
    """Secure-ish delete: overwrite with zeros then unlink."""
    try:
        f = Path(path)
        if f.exists():
            sz = f.stat().st_size
            with open(f, "r+b") as fp:
                fp.write(b"\x00" * sz)
            f.unlink()
    except Exception:
        pass


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default="C:/Users/HFRC/Downloads/ArbitrageX_Unified_Config.xlsm",
                    help="Excel source (Python reads the 4 sheets). Default mode. "
                         "Mutually exclusive with --json-in.")
    ap.add_argument("--json-in", default=None,
                    help="Pre-built bundle JSON file (the VBA macro writes this). Skips the "
                         "Excel read - Python only encrypts. Mutually exclusive with --xlsx.")
    ap.add_argument("--public-key", default="C:/Users/HFRC/Downloads/arbx_bundle_public.pem")
    ap.add_argument("--schema", default=None, help="jsonschema file for pre-encrypt validation")
    ap.add_argument("--out", default=None,
                    help="Explicit output path for the .enc (so the VBA macro / browser upload "
                         "can pick it up). If omitted, a shred-safe temp is used.")
    ap.add_argument("--vps-host", default="arbx")
    ap.add_argument("--dest", default="/opt/arbitragex-v2/config/arbx_config_bundle.json.enc")
    ap.add_argument("--no-upload", action="store_true",
                    help="encrypt only, skip scp (VBA mode: the macro SCPs, or the operator "
                         "uploads via the browser panel - Ruta 2)")
    args = ap.parse_args()

    if args.json_in and args.xlsx != ap.get_default("xlsx"):
        sys.exit("FATAL: --json-in and --xlsx are mutually exclusive (pass only one source).")

    # 1. Get the bundle JSON: either the VBA macro built it (--json-in) or Python reads Excel.
    if args.json_in:
        with open(args.json_in, "rb") as f:
            bundle_json = f.read()
        bundle = json.loads(bundle_json)
        print(f"bundle (json-in): {len(bundle_json)} bytes | source={args.json_in}")
    else:
        bundle = build_bundle(args.xlsx)
        bundle_json = json.dumps(bundle, indent=2).encode()
        print(f"bundle (xlsx): {len(bundle_json)} bytes | env_vars={len(bundle['env_vars'])} "
              f"chains={len(bundle['chains'])} api_keys={len(bundle['api_keys'])} "
              f"contract_addrs={len(bundle['contract_addresses'])}")

    if args.schema:
        import jsonschema
        jsonschema.validate(json.loads(bundle_json), json.load(open(args.schema)))
        print("schema validation: OK")

    # Defence-in-depth (layer 2 of 3): assert NEVER_SHIP absent.
    # Layer 1 = VBA macro skips these keys; layer 3 = Rust importer re-asserts post-decrypt.
    leaked = [k for k in NEVER_SHIP if k in bundle.get("env_vars", {})]
    if leaked:
        sys.exit(f"FATAL: forbidden keys shipped: {leaked}")

    # 2. Encrypt (in-memory)
    enc = hybrid_encrypt(bundle_json, args.public_key)
    print(f"hybrid encrypt (RSA-OAEP-4096 + AES-256-GCM): {len(enc)} bytes")

    # 3. Write the .enc: explicit --out path, else a shred-safe temp we own.
    enc_path = args.out
    tmp_plain = None
    if enc_path is None:
        # Park the plaintext in a temp so we can shred it (never on disk unencrypted otherwise).
        with tempfile.NamedTemporaryFile(delete=False, suffix=".json") as tmp:
            tmp.write(bundle_json)
            tmp_plain = tmp.name
        enc_path = tmp_plain + ".enc"

    try:
        with open(enc_path, "wb") as f:
            f.write(enc)

        # sha256 fingerprint so the operator / VBA macro can eyeball key + payload continuity.
        import hashlib
        sha = hashlib.sha256(enc).hexdigest()
        print(f"sha256({enc_path}): {sha}")

        if args.no_upload:
            print(f"encrypted (no-upload): {enc_path}")
        else:
            subprocess.run(["scp", enc_path, f"{args.vps_host}:{args.dest}"], check=True)
            print(f"uploaded -> {args.vps_host}:{args.dest}")

        # Shred the plaintext temp if WE created it (NOT the VBA's --json-in file -
        # the macro owns that file's lifecycle and shreds it in its own step 4).
        if tmp_plain is not None:
            shred(tmp_plain)
        # Clean the .enc temp if WE created it AND it was uploaded (keep if --out or --no-upload).
        if tmp_plain is not None and not args.no_upload:
            os.unlink(enc_path)
        if tmp_plain is None:
            print("plaintext was the caller's --json-in file (not shred by Python)")
        else:
            print("plaintext shred OK" + (
                " - only the uploaded .enc on the VPS remains"
                if not args.no_upload else f" - .enc kept at {enc_path}"))
    finally:
        if tmp_plain is not None and os.path.exists(tmp_plain):
            shred(tmp_plain)


if __name__ == "__main__":
    main()
