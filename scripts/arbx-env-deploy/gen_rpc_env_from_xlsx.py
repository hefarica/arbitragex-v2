#!/usr/bin/env python3
"""
Generate RPC_HTTP_<chainId> / RPC_WS_<chainId> .env values from the
ArbitrageX_Unified_Config.xlsm catalog.

Direction: Excel = SSOT, ONE-WAY (Excel -> .env). Read-only on the workbook.

Sources:
  - "RPC Providers"  : Chain | Protocolo (HTTP/WSS) | Proveedor | URL
  - "_RED_lookup"    : Proveedor -> token (PROVEEDOR)   (e.g. PublicNode -> publicnode)

Emits, per MAINNET chain, the multi-provider lists in the EXACT .env format the
hot-path already consumes:
  RPC_HTTP_1=publicnode=https://...,drpc=https://...,lava=https://...
  RPC_WS_1=publicnode=wss://...,drpc=wss://...

Only complete, keyless public endpoints are emitted (key-requiring URLs are
skipped unless --include-keyed). No secrets are read or written. Output goes to
stdout AND a gitignored fragment (rpc_env_generated.env) for review; applying it
to the '.env Production' sheet / VPS .env is a separate, operator-gated step.
"""
import argparse
import os
import sys
from collections import OrderedDict, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# Standard public chain IDs for the MAINNETS the catalog covers. These are
# universal protocol constants (chainlist.org), not operator-specific config.
# Testnets are intentionally omitted: the .env handles them via dedicated
# single-URL vars (SEPOLIA_RPC_URL, ARBITRUM_SEPOLIA_RPC_URL, HOLESKY_RPC_URL).
CHAIN_IDS = {
    "Ethereum Mainnet": 1,
    "Optimism": 10,
    "BSC Mainnet": 56,
    "Gnosis": 100,
    "Polygon Mainnet": 137,
    "Base": 8453,
    "Arbitrum One": 42161,
    "Avalanche": 43114,
    "Linea": 59144,
    "Blast": 81457,
    "Scroll": 534352,
}
PLACEHOLDER_MARKERS = ("<", ">", "${", "YOUR", "API_KEY", "apikey", "key=")
# '/v2/' was removed from PLACEHOLDER_MARKERS because it also matches REAL
# Alchemy keys (/v2/alch_...). Real keys start with 'alch_' and are >20 chars.
# We keep the /v2/ skip ONLY when the segment after /v2/ looks like a placeholder.

# ARBX-R-0003: providers VERIFIED DEAD on prod (MC-RPC-1 provider-breakdown
# evidence, 2026-08-14 boot audit) and therefore EXCLUDED from every generated
# list even though the workbook catalog still lists them. The Excel remains the
# SSOT for what EXISTS; this denylist is the evidence-backed layer for what is
# USABLE. Remove an entry here only with fresh contrary evidence (and update the
# workbook in the same operator pass).
DEAD_PROVIDERS = {
    "llama": "Cloudflare challenge HTML instead of JSON-RPC (eth.llamarpc.com)",
    "0xrpc": "404 — endpoint retired",
    "1rpc": "403 Forbidden on boot eth_chainid",
}


def _is_placeholder_url(url: str) -> bool:
    """True when the URL has placeholder markers OR a /v2/<placeholder> segment."""
    u = url.strip()
    if any(m in u for m in PLACEHOLDER_MARKERS):
        return True
    if "/v2/" in u:
        key_part = u.split("/v2/")[1].split("/")[0].split("?")[0]
        # Real Alchemy keys: alch_ + 30+ chars. Real Infura: 32 hex chars.
        if len(key_part) >= 20 and (key_part.startswith("alch_") or key_part.startswith("0x") or len(key_part) == 32):
            return False  # Looks like a real key — include it
        # Short/generic segment after /v2/ = placeholder
        if len(key_part) < 20 or key_part.upper() in ("YOUR_KEY", "YOURKEY", "APIKEY", "KEY", "SECRET"):
            return True
    return False


def provider_token_map(wb):
    ws = wb["_RED_lookup"]
    m = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        tok = r[6] if len(r) > 6 else None
        name = r[7] if len(r) > 7 else None
        if name and tok and name not in m:
            m[str(name).strip()] = str(tok).strip()
    return m


def main():
    ap = argparse.ArgumentParser(description="Generate RPC_HTTP_/RPC_WS_ .env values from the config workbook (Excel=SSOT).")
    ap.add_argument("--xlsx", default=os.environ.get("ARBX_CONFIG_XLSX", r"C:\Users\HFRC\Downloads\ArbitrageX_Unified_Config.xlsm"))
    ap.add_argument("--out", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "rpc_env_generated.env"))
    ap.add_argument("--include-keyed", action="store_true", help="include providers whose URL looks key-requiring (default: skip)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    provmap = provider_token_map(wb)
    ws = wb["RPC Providers"]

    acc = defaultdict(lambda: {"HTTP": OrderedDict(), "WS": OrderedDict()})
    unknown_chains, skipped, dead_excluded = set(), [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        chain = r[0]
        proto = r[1]
        prov = r[2]
        url = r[3] if len(r) > 3 else None
        if not chain or not proto or not url:
            continue
        cid = CHAIN_IDS.get(str(chain).strip())
        if cid is None:
            unknown_chains.add(str(chain).strip())
            continue
        u = str(url).strip().rstrip("/")
        if (not args.include_keyed) and _is_placeholder_url(u):
            skipped.append(f"{chain} | {proto} | {prov} | {u}")
            continue
        pkey = "WS" if "WS" in str(proto).upper() else "HTTP"
        tok = provmap.get(str(prov).strip()) or str(prov).strip().lower().replace(" ", "")
        # ARBX-R-0003: dead-provider denylist (evidence in DEAD_PROVIDERS) —
        # excluded from the generated list regardless of catalog presence.
        if tok in DEAD_PROVIDERS:
            dead_excluded.append(f"{chain} | {pkey} | {tok} | {DEAD_PROVIDERS[tok]}")
            continue
        acc[cid][pkey].setdefault(tok, u)  # first occurrence wins (catalog order = priority)
    wb.close()

    lines = []
    for cid in sorted(acc.keys()):
        if acc[cid]["HTTP"]:
            lines.append(f"RPC_HTTP_{cid}=" + ",".join(f"{t}={u}" for t, u in acc[cid]["HTTP"].items()))
        if acc[cid]["WS"]:
            lines.append(f"RPC_WS_{cid}=" + ",".join(f"{t}={u}" for t, u in acc[cid]["WS"].items()))

    header = (
        "# AUTO-GENERATED from ArbitrageX_Unified_Config.xlsm (RPC Providers + _RED_lookup).\n"
        "# Excel = SSOT, one-way. Regenerate: python scripts/arbx-env-deploy/gen_rpc_env_from_xlsx.py\n"
        "# Do NOT hand-edit; edit the workbook catalog and regenerate.\n"
        + (
            "# ARBX-R-0003: dead providers excluded with evidence: "
            + "; ".join(f"{k} ({v})" for k, v in DEAD_PROVIDERS.items())
            + "\n"
            if DEAD_PROVIDERS
            else ""
        )
    )
    with open(args.out, "w", encoding="utf-8", newline="\n") as f:
        f.write(header)
        f.write("\n".join(lines) + "\n")

    print(f"Wrote {len(lines)} env lines for {len(acc)} mainnet chains -> {args.out}\n")
    print("\n".join(lines))
    if unknown_chains:
        print("\n# Skipped (no mainnet chainId mapping — testnets use dedicated *_RPC_URL vars):")
        for c in sorted(unknown_chains):
            print(f"#   {c}")
    if skipped:
        print(f"\n# Skipped {len(skipped)} key-requiring/placeholder URL(s) (use --include-keyed to force):")
        for s in skipped:
            print(f"#   {s}")
    if dead_excluded:
        print(f"\n# ARBX-R-0003: EXCLUDED {len(dead_excluded)} dead-provider row(s) (see DEAD_PROVIDERS):")
        for s in dead_excluded:
            print(f"#   {s}")


if __name__ == "__main__":
    main()
