"""Ingesta del workbook ArbitrageX_Dynamic_QuoteBase_Route_Manual_264.xlsx.

Pase 1: manifest estructural (dims, celdas, formulas, headers) -> quotebase_manifest.json
Pase 2: dumps profundos de hojas críticas -> quotebase_deep_dump.txt

Mismo estilo que docs/excel_ingestion_manifest.json (programa XLS-* previo).
"""
import json
import os
import re
from pathlib import Path

import openpyxl

SRC = Path(os.environ.get(
    "ARBX_QUOTEBASE_XLSX",
    r"C:\Users\HFRC\Downloads\ArbitrageX_Dynamic_QuoteBase_Route_Manual_264.xlsx",
))
OUT_DIR = Path(__file__).parent


def cell_repr(v):
    if v is None:
        return ""
    s = str(v)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:80]


def main():
    wb_val = openpyxl.load_workbook(SRC, data_only=True)
    wb_frm = openpyxl.load_workbook(SRC, data_only=False)

    manifest = {"path": str(SRC), "total_sheets": len(wb_val.worksheets), "sheets": {}}
    deep = []

    for ws in wb_val.worksheets:
        ws_f = wb_frm[ws.title]
        non_empty = 0
        formulas = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    non_empty += 1
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1

        headers = []
        for r in range(1, 4):
            headers.append([cell_repr(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)])

        manifest["sheets"][ws.title] = {
            "state": ws.sheet_state,
            "dimensions": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "non_empty_cells": non_empty,
            "formulas": formulas,
            "header_rows": headers,
        }

    OUT_DIR.joinpath("quotebase_manifest.json").write_text(
        json.dumps(manifest, indent=1, ensure_ascii=False), encoding="utf-8"
    )

    # ---- dumps profundos de hojas críticas ----
    def dump_sheet(name, max_rows=None, max_cols=None):
        ws = wb_val[name]
        mr = max_rows or ws.max_row
        mc = max_cols or ws.max_column
        deep.append(f"\n{'='*100}\nSHEET {name}  ({ws.max_row}x{ws.max_column})  mostrando {mr}x{mc}\n{'='*100}")
        for r in range(1, mr + 1):
            vals = [cell_repr(ws.cell(row=r, column=c).value) for c in range(1, mc + 1)]
            if any(vals):
                deep.append(f"r{r:>4} | " + " | ".join(vals))

    dump_sheet("00_MANUAL")
    dump_sheet("01_CONFIG")
    dump_sheet("02_ALLOWED_SYMBOLS", max_rows=12)
    dump_sheet("03_CHAIN_REGISTRY", max_rows=8)
    dump_sheet("04_INDEX_MATH")
    dump_sheet("05_QUOTE_BASE")
    dump_sheet("06_EDGE_MATH")
    dump_sheet("07_INEFFICIENCY", max_rows=20)
    dump_sheet("08_HOPS_2_7")
    dump_sheet("09_RUNTIME_STRUCTURES")
    dump_sheet("10_LATENCY")
    dump_sheet("11_STRATEGY_HOP_MAP", max_rows=6)
    dump_sheet("12_STRATEGY_HOP_EXPANDED", max_rows=6)
    dump_sheet("13_DETECTOR_POLICY", max_rows=8)
    dump_sheet("14_RESEARCH")
    dump_sheet("15_IMPLEMENTATION_CONTRACT")
    dump_sheet("16_COVERAGE")

    OUT_DIR.joinpath("quotebase_deep_dump.txt").write_text("\n".join(deep), encoding="utf-8")
    print(f"manifest -> {OUT_DIR / 'quotebase_manifest.json'}")
    print(f"deep dump -> {OUT_DIR / 'quotebase_deep_dump.txt'} ({len(deep)} lineas)")
    for name, s in manifest["sheets"].items():
        print(f"{name:32s} {s['max_row']:>5}x{s['max_col']:<3} cells={s['non_empty_cells']:>6} formulas={s['formulas']}")


if __name__ == "__main__":
    main()
