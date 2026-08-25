"""Extracción canónica de datos del workbook QUOTEBASE-264 + validación diferencial.

Outputs (docs/, ubicación canónica — gen_hopmask_rs.py consume de ahí):
- quotebase_config.json            knobs de 01_CONFIG (count DERIVADO del Excel, sin hardcode)
- quotebase_strategy_hop_map.json  264 estrategias × 27 cols (11_STRATEGY_HOP_MAP)
- quotebase_strategy_hop_expanded.json  1.436 combos × 18 cols (12_STRATEGY_HOP_EXPANDED)
- quotebase_detector_policy.json   60 detectores (13_DETECTOR_POLICY)
- quotebase_research.json          8 referencias (14_RESEARCH)

Validación diferencial (Excel ↔ código):
- PairIndex(i,j,N) = i(2N-i-1)/2 + (j-i-1)  vs ejemplo trabajado del 04_INDEX_MATH (i=3,j=17,N=22 → 73)
- HopMask_u8 recomputado desde H2..H7 vs columna HopMask_u8  (bit k ⇔ hop k+2)
- Distribución por hop vs 16_COVERAGE: 245/262/260/233/233/203, total 1436
- Inyectividad de PairIndex para N=22 (231 pares, 0 colisiones)
"""
import json
import os
from pathlib import Path

import openpyxl

SRC = Path(os.environ.get(
    "ARBX_QUOTEBASE_XLSX",
    r"C:\Users\HFRC\Downloads\ArbitrageX_Dynamic_QuoteBase_Route_Manual_264.xlsx",
))
OUT = Path(__file__).resolve().parents[2] / "docs"
checks = []


def check(name, ok, detail=""):
    checks.append({"check": name, "pass": bool(ok), "detail": detail})


def sheet_rows(ws, header_row=3, min_nonempty=1):
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if sum(v is not None and str(v).strip() != "" for v in vals) >= min_nonempty:
            rows.append({h: v for h, v in zip(headers, vals) if h})
    return rows


def pair_index(i, j, n):
    a, b = (i, j) if i < j else (j, i)
    return a * (2 * n - a - 1) // 2 + (b - a - 1)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # --- 01_CONFIG knobs ---
    # Diferencial SIN hardcode: el count esperado se DERIVA del propio Excel canónico.
    # Fila de parámetro = fila con Parameter O Value no-vacío en la región de datos (4..max_row).
    # Si el workbook cambia de 17 a N knobs, el check sigue siendo válido (actual == canonical).
    ws = wb["01_CONFIG"]
    knobs = []
    canonical_rows = 0
    malformed_rows = []
    for r in range(4, ws.max_row + 1):
        p = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        has_p = p is not None and str(p).strip() != ""
        has_v = v is not None and str(v).strip() != ""
        if has_p or has_v:
            canonical_rows += 1
        if has_p and not has_v:
            malformed_rows.append(r)  # parámetro sin valor — visible, no silencioso
        if has_p:
            knobs.append({
                "parameter": str(p).strip(),
                "value": v,
                "unit": ws.cell(row=r, column=3).value,
                "meaning": ws.cell(row=r, column=4).value,
                "runtime_binding": ws.cell(row=r, column=5).value,
            })
    (OUT / "quotebase_config.json").write_text(json.dumps(knobs, indent=1, ensure_ascii=False), encoding="utf-8")
    check("01_CONFIG knobs == canonical_excel_count (derivado de la hoja, sin hardcode)",
          len(knobs) == canonical_rows,
          f"actual={len(knobs)} canonical={canonical_rows}")
    from collections import Counter
    dup = [k for k, n in Counter(x["parameter"] for x in knobs).items() if n > 1]
    check("01_CONFIG knobs sin parámetros duplicados", not dup, f"duplicados: {dup}")
    check("01_CONFIG knobs sin parámetro-valor ausente", not malformed_rows,
          f"filas con Parameter pero sin Value: {malformed_rows}")

    # --- 11_STRATEGY_HOP_MAP (264) ---
    ws = wb["11_STRATEGY_HOP_MAP"]
    strategies = sheet_rows(ws)
    (OUT / "quotebase_strategy_hop_map.json").write_text(
        json.dumps(strategies, indent=1, ensure_ascii=False), encoding="utf-8")
    check("11 hop_map strategies == 264", len(strategies) == 264, f"got {len(strategies)}")

    # HopMask diferencial: bit(h-2) por h en 2..7
    mask_mismatch = []
    for s in strategies:
        bits = 0
        for h in range(2, 8):
            v = s.get(f"H{h}")
            if v is True or str(v).strip().upper() == "TRUE":
                bits |= 1 << (h - 2)
        if int(s.get("HopMask_u8") or 0) != bits:
            mask_mismatch.append(s.get("MEV_ID"))
    check("HopMask_u8 == recomputado(H2..H7) en 264/264", not mask_mismatch, f"mismatches: {mask_mismatch[:5]}")

    # Distribución por hop vs 16_COVERAGE
    expected = {2: 245, 3: 262, 4: 260, 5: 233, 6: 233, 7: 203}
    for h, exp in expected.items():
        got = sum(1 for s in strategies if str(s.get(f"H{h}")).strip().upper() == "TRUE")
        check(f"hop {h} compatibles == {exp}", got == exp, f"got {got}")

    # --- 12_STRATEGY_HOP_EXPANDED (1436) ---
    ws = wb["12_STRATEGY_HOP_EXPANDED"]
    combos = sheet_rows(ws)
    (OUT / "quotebase_strategy_hop_expanded.json").write_text(
        json.dumps(combos, indent=1, ensure_ascii=False), encoding="utf-8")
    check("12 combos == 1436", len(combos) == 1436, f"got {len(combos)}")
    check("12 combos == suma(245+262+260+233+233+203)=1436",
          len(combos) == sum(expected.values()))

    # Cobertura bidireccional: cada combo (MEV_ID,hop) corresponde a un True del mapa
    map_true = {(s["MEV_ID"], h) for s in strategies for h in range(2, 8)
                if str(s.get(f"H{h}")).strip().upper() == "TRUE"}
    combo_set = {(c["MEV_ID"], int(c["Hop"])) for c in combos}
    check("mapa↔expandido biyectivo", map_true == combo_set,
          f"solo_en_mapa={len(map_true - combo_set)} solo_en_expandido={len(combo_set - map_true)}")

    # --- 13_DETECTOR_POLICY (60) ---
    ws = wb["13_DETECTOR_POLICY"]
    detectors = sheet_rows(ws)
    (OUT / "quotebase_detector_policy.json").write_text(
        json.dumps(detectors, indent=1, ensure_ascii=False), encoding="utf-8")
    check("13 detectores == 60", len(detectors) == 60, f"got {len(detectors)}")

    # Detectores del mapa ⊆ política de detectores
    det_ids = {d["Detector_ID"] for d in detectors}
    map_dets = {s["Detector_ID"] for s in strategies}
    check("Detector_IDs del mapa ⊆ política 60", map_dets <= det_ids,
          f"sin política: {sorted(map_dets - det_ids)[:5]}")

    # --- 14_RESEARCH (8) ---
    ws = wb["14_RESEARCH"]
    refs = sheet_rows(ws)
    (OUT / "quotebase_research.json").write_text(
        json.dumps(refs, indent=1, ensure_ascii=False), encoding="utf-8")
    check("14 research refs == 8", len(refs) == 8, f"got {len(refs)}")

    # --- PairIndex diferencial vs 04_INDEX_MATH ejemplo (i=3, j=17, N=22 → 73) ---
    got = pair_index(3, 17, 22)
    check("PairIndex(3,17,22) == 73 (ejemplo workbook)", got == 73, f"got {got}")

    # Inyectividad N=22
    seen = {}
    collisions = 0
    for i in range(22):
        for j in range(i + 1, 22):
            k = pair_index(i, j, 22)
            if k in seen:
                collisions += 1
            seen[k] = (i, j)
    check("PairIndex inyectivo N=22 (231 pares, 0 colisiones)",
          collisions == 0 and len(seen) == 231, f"colisiones={collisions} pares={len(seen)}")
    check("PairIndex rango [0, C(22,2)-1=230]",
          min(seen) == 0 and max(seen) == 230, f"min={min(seen)} max={max(seen)}")

    # --- resumen de superficies y graph models ---
    from collections import Counter
    surfaces = Counter(s["Surface"] for s in strategies)
    graphs = Counter(s["Graph_Model"] for s in strategies)
    print("SUPERFICIES:", dict(surfaces))
    print("GRAPH_MODELS:", dict(graphs))
    print("STATUS:", dict(Counter(s.get("Status") for s in strategies)))
    print("EXECUTION_CLASS:", dict(Counter(s.get("Execution_Class") for s in strategies)))

    (OUT / "quotebase_extraction_checks.json").write_text(
        json.dumps(checks, indent=1, ensure_ascii=False), encoding="utf-8")
    failed = [c for c in checks if not c["pass"]]
    print(f"\nCHECKS: {len(checks) - len(failed)}/{len(checks)} PASS")
    for c in failed:
        print(f"  FAIL: {c['check']} — {c['detail']}")


if __name__ == "__main__":
    main()
